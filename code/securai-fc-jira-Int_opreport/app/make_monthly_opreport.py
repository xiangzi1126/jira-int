import os
import re
import glob
import zipfile
import shutil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.plotarea import DataTable
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import cm_to_EMU
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import logging

# --- I. 常量定义 ---
PROJECT_KEY = 'TSIINFRA'

SHEET_COVER = '表紙'
SHEET_SUMMARY = '総評1-8'
SHEET_CPU = 'リソース (CPU)'
SHEET_MEMORY = 'リソース (Memory)'
SHEET_DISK = 'リソース (Disk )'
SHEET_TREND = '推移'

# 表紙 常量
COVER_COMPANY_ROW = 3
COVER_SYSTEM_ROW = 8
COVER_DATE_ROW = 29

# 推移 常量
TREND_HEADER_ROW = 2
TREND_DATA_START_ROW = 3

# リソース 常量
RESOURCE_TITLE_START_ROW = 3
RESOURCE_INTERVAL_ROWS = 18
RESOURCE_DATA_START_COL = 20  # T列

# 総評 布局配置
LAYOUT = {
    'greeting_row': 2,
    'summary_start_row': 5,  # 1、総評 概要文本起始行
    'alert_count_row': 8,  # 2-1アラート件数（匹配模板）
    'dynamic_clear_from': 42,  # 从此行起清空并重写
    'chart_anchor': 'A10',
}


# --- II. 辅助函数 ---
def init_logger():
    log_dir = os.path.normpath(os.path.join(os.path.dirname(__file__), '../../jira/log'))
    os.makedirs(log_dir, exist_ok=True)
    log_file = os.path.join(log_dir, 'monthly_opreport_TSIINFRA.log')
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler()
        ]
    )
    return logging.getLogger(__name__)


logger = init_logger()


def safe_read_csv(file_path):
    """多编码尝试读取CSV"""
    for encoding in ['utf-8-sig', 'utf-8', 'gbk', 'shift-jis', 'latin1']:
        try:
            return pd.read_csv(file_path, encoding=encoding)
        except (UnicodeDecodeError, UnicodeError):
            continue
        except Exception as e:
            logger.warning(f"读取文件 {file_path} 失败: {e}")
            continue
    raise IOError(f"无法读取文件: {file_path}")


def parse_account_tag(tag_str):
    """将 'key:TSI value:infra' 转换为 'TSI:infra'"""
    if not tag_str or pd.isna(tag_str) or str(tag_str).strip() in ('', 'nan'):
        return None
    tag_str = str(tag_str).strip()
    key_match = re.search(r'key:(\S+)', tag_str)
    val_match = re.search(r'value:(\S+)', tag_str)
    if key_match and val_match:
        return f"{key_match.group(1)}:{val_match.group(1)}"
    return tag_str


def get_alert_circle_number(n):
    """数字转圆圈数字 ①②③..."""
    circles = '①②③④⑤⑥⑦⑧⑨⑩⑪⑫⑬⑭⑮⑯⑰⑱⑲⑳'
    if 1 <= n <= 20:
        return circles[n - 1]
    return f'({n})'


def _set_chart_size(chart_obj, anchor_str, width_cm, height_cm):
    """设置图表尺寸（内存中设置，保存后用 _fix_chart_sizes_in_xlsx 修正 XML）"""
    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
    col_letter, row = coordinate_from_string(anchor_str)
    col = column_index_from_string(col_letter)
    marker = AnchorMarker(col=col - 1, colOff=0, row=row - 1, rowOff=0)
    size = XDRPositiveSize2D(cx=cm_to_EMU(width_cm), cy=cm_to_EMU(height_cm))
    chart_obj.anchor = OneCellAnchor(_from=marker, ext=size)


def _fix_chart_sizes_in_xlsx(xlsx_path, size_map):
    """保存后修正 xlsx 内 drawing XML 的图表尺寸
    size_map: {drawing序号(1-based): (width_cm, height_cm)}
    对该 drawing 内的所有 ext 统一替换
    """
    tmp_path = xlsx_path + '.tmp'
    pat = re.compile(r'(<ext cx=")([0-9]+)(" cy=")([0-9]+)(")')

    with zipfile.ZipFile(xlsx_path, 'r') as zin:
        with zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                m = re.match(r'xl/drawings/drawing(\d+)\.xml', item.filename)
                if m:
                    drawing_idx = int(m.group(1))
                    if drawing_idx in size_map:
                        w_cm, h_cm = size_map[drawing_idx]
                        new_cx = str(cm_to_EMU(w_cm))
                        new_cy = str(cm_to_EMU(h_cm))
                        text = data.decode('utf-8')
                        text = pat.sub(
                            lambda g: g.group(1) + new_cx + g.group(3) + new_cy + g.group(5),
                            text
                        )
                        data = text.encode('utf-8')
                zout.writestr(item, data)

    # Windows 安全替换
    os.remove(xlsx_path)
    shutil.move(tmp_path, xlsx_path)
    logger.info(f"图表尺寸修正完成: {size_map}")


# --- III. 数据加载 ---
def load_all_data(script_dir):
    """加载所有数据源"""

    def get_path(rel_path):
        return os.path.normpath(os.path.join(script_dir, rel_path))

    data_dir = get_path("../../jira/data")
    output_dir = get_path("../../jira/monthly_report")
    template_path = get_path(
        "../../jira/monthly_report/运维月报模板_【株式会社TSI様向け】インフラ月次報告書.xlsx"
    )
    os.makedirs(output_dir, exist_ok=True)

    # 数据文件
    op_details_csv = os.path.join(data_dir, 'op_resale_business_details.csv')
    op_account_csv = os.path.join(data_dir, 'jira_get_op_account.csv')
    ecs_csv = os.path.join(data_dir, 'aliyun_ecs_op.csv')
    cmd_summary_csv = os.path.join(data_dir, 'aliyun_op_cmd_summary.csv')
    alert_csv = os.path.join(data_dir, 'jira_get_alert.csv')

    for path in [op_details_csv, op_account_csv, ecs_csv, cmd_summary_csv, alert_csv]:
        if not os.path.exists(path):
            raise FileNotFoundError(f"关键文件不存在: {path}")

    if not os.path.exists(template_path):
        raise FileNotFoundError(f"模板文件不存在: {template_path}")

    op_details_df = safe_read_csv(op_details_csv)
    op_account_df = safe_read_csv(op_account_csv)
    ecs_df = safe_read_csv(ecs_csv)
    cmd_summary_df = safe_read_csv(cmd_summary_csv)
    alert_df = safe_read_csv(alert_csv)

    # 配置文件
    names_csv = get_path("../../jira/config/opreport_names.csv")
    hosts_csv = get_path("../../jira/config/opreport_hosts.csv")
    names_df = safe_read_csv(names_csv) if os.path.exists(names_csv) else pd.DataFrame()
    hosts_df = safe_read_csv(hosts_csv) if os.path.exists(hosts_csv) else pd.DataFrame()

    logger.info(f"数据加载完成: 项目{len(op_details_df)}个, 账号{len(op_account_df)}个, "
                f"ECS{len(ecs_df)}台, 资源汇总{len(cmd_summary_df)}条, 告警{len(alert_df)}条")
    if not names_df.empty:
        logger.info(f"名称映射: {len(names_df)}条, 主机映射: {len(hosts_df)}条")

    return {
        'output_dir': output_dir,
        'template_path': template_path,
        'data_dir': data_dir,
        'op_details': op_details_df,
        'op_account': op_account_df,
        'ecs': ecs_df,
        'cmd_summary': cmd_summary_df,
        'alert': alert_df,
        'names': names_df,
        'hosts': hosts_df,
    }


def get_project_instances(account, tag_filter, ecs_df):
    """根据账号和标签过滤ECS实例"""
    account_df = ecs_df[ecs_df['资源所属账号'] == account]
    if tag_filter:
        tag_mask = account_df['标签'].astype(str).str.contains(re.escape(tag_filter), na=False)
        account_df = account_df[tag_mask]
    return account_df


def get_alerts_for_project(alert_df, project_key, target_year, target_month):
    """获取项目的今月和前月告警（排除 备注='不记录月报'）"""
    proj_alerts = alert_df[alert_df['Project Key'] == project_key].copy()
    if proj_alerts.empty:
        return pd.DataFrame(), pd.DataFrame()

    # 过滤：排除 备注="不记录月报"
    if '备注' in proj_alerts.columns:
        proj_alerts = proj_alerts[
            proj_alerts['备注'].astype(str).str.strip() != '不记录月报'
            ]

    if proj_alerts.empty:
        return pd.DataFrame(), pd.DataFrame()

    # 解析告警检测日时（去除时区）
    proj_alerts['告警检测日时'] = pd.to_datetime(
        proj_alerts['告警检测日时'], errors='coerce', utc=True
    )
    proj_alerts['告警检测日时'] = proj_alerts['告警检测日时'].dt.tz_localize(None)
    proj_alerts = proj_alerts.dropna(subset=['告警检测日时'])

    # 今月/前月范围
    current_start = datetime(target_year, target_month, 1)
    current_end = current_start + relativedelta(months=1) - timedelta(seconds=1)
    prev_start = current_start - relativedelta(months=1)
    prev_end = current_start - timedelta(seconds=1)

    current_alerts = proj_alerts[
        (proj_alerts['告警检测日时'] >= current_start) &
        (proj_alerts['告警检测日时'] <= current_end)
        ].sort_values('告警检测日时')
    prev_alerts = proj_alerts[
        (proj_alerts['告警检测日时'] >= prev_start) &
        (proj_alerts['告警检测日时'] <= prev_end)
        ]
    return current_alerts, prev_alerts


def load_monitoring_data(data_dir, account, instance_name, metric_type):
    """加载单个主机的监控时序数据"""
    monitor_dir = os.path.join(data_dir, f'监控数据_{account}')
    if not os.path.exists(monitor_dir):
        return pd.DataFrame()

    pattern = os.path.join(monitor_dir, f'{metric_type}_*.csv')
    files = glob.glob(pattern)

    target_file = None
    for f in files:
        basename = os.path.basename(f)
        name_part = basename.replace(f'{metric_type}_', '').replace('.csv', '')
        if name_part.lower() == instance_name.lower():
            target_file = f
            break

    if not target_file:
        return pd.DataFrame()

    try:
        df = safe_read_csv(target_file)
        if 'timestamp' in df.columns:
            df['datetime'] = pd.to_datetime(df['timestamp'], unit='ms')
            df = df.sort_values('datetime')
        return df
    except Exception as e:
        logger.warning(f"读取监控数据失败 {target_file}: {e}")
        return pd.DataFrame()


def get_host_mapping(hosts_df, project_key):
    """获取项目的主机映射表（显示名、环境）"""
    if hosts_df.empty:
        return {}
    proj_hosts = hosts_df[hosts_df['Project Key'] == project_key]
    mapping = {}
    for _, row in proj_hosts.iterrows():
        mapping[str(row['InstanceId'])] = {
            'display_name': row['显示名'],
            'env': row['环境'],
        }
    return mapping


# --- IV. Sheet 填充函数 ---
def populate_cover(wb, company_name, system_name, target_year, target_month, today):
    """填充表紙"""
    if SHEET_COVER not in wb.sheetnames:
        return
    ws = wb[SHEET_COVER]
    ws.cell(row=COVER_COMPANY_ROW, column=1).value = f"{company_name}\u3000御中"
    ws.cell(row=COVER_SYSTEM_ROW, column=1).value = (
        f"{system_name}\n{target_year}年{target_month}月\u3000月次報告"
    )
    ws.cell(row=COVER_DATE_ROW, column=1).value = (
        f"\n作成日：{today.year}/{today.month}/{today.day}\nセキュライ大連"
    )
    logger.info(f"【表紙】填充完成: {company_name}, {target_year}年{target_month}月")


def populate_summary(wb, current_alerts, prev_alerts, cmd_data,
                     target_year, target_month, host_mapping):
    """填充総評1-8（动态布局：R42以下清空重写）"""
    if SHEET_SUMMARY not in wb.sheetnames:
        return
    ws = wb[SHEET_SUMMARY]

    # R2: 月份挨拶
    ws.cell(row=LAYOUT['greeting_row'], column=1).value = (
        f"{target_year}年{target_month}月のシステム稼働状況について、下記の通り、ご報告申し上げます。"
    )

    alert_count = len(current_alerts)

    # 1、総評 概要（来自 备注 列）
    if alert_count > 0:
        for idx, (_, row) in enumerate(current_alerts.iterrows()):
            detect_time = row['告警检测日时']
            remark = row.get('备注', '')
            if pd.isna(remark) or not str(remark).strip():
                remark = f"{row.get('告警目标', '')}サーバのアラートを検知しました。"
            date_str = f"{detect_time.month}/{detect_time.day}"
            ws.cell(row=LAYOUT['summary_start_row'] + idx, column=1).value = (
                f"{date_str} 、{str(remark).strip()}"
            )

    # 2、アラート対応状況
    ws.cell(row=LAYOUT['alert_count_row'] - 1, column=1).value = "2、アラート対応状況"

    # 2-1アラート件数
    ws.cell(row=LAYOUT['alert_count_row'], column=1).value = f"2-1アラート件数：{alert_count}件"

    # --- 动态区域：清空 R42 以下所有内容和合并单元格 ---
    clear_from = LAYOUT['dynamic_clear_from']
    _clear_dynamic_area(ws, clear_from)

    # 动态写入 section headers + 数据
    current_row = clear_from
    ws.cell(row=current_row, column=1).value = '2-2説明'
    current_row += 1

    # ■検知内容
    ws.cell(row=current_row, column=1).value = '■検知内容'
    current_row += 1

    if alert_count > 0:
        for idx, (_, row) in enumerate(current_alerts.iterrows()):
            detect_time = row['告警检测日时']
            alert_desc = row.get('告警简述', '')
            if pd.isna(alert_desc):
                alert_desc = ''
            alert_desc = str(alert_desc).strip()
            date_str = (f"{detect_time.year}-{detect_time.month}-{detect_time.day} "
                        f"{detect_time.hour}:{detect_time.minute:02d}")
            circle_num = get_alert_circle_number(idx + 1)
            text = f"アラート{circle_num}：{date_str} {alert_desc}"
            ws.cell(row=current_row, column=1).value = text
            current_row += 1

    # 空一行
    current_row += 1

    # ■事象纏め
    ws.cell(row=current_row, column=1).value = '■事象纏め'
    current_row += 1

    if alert_count > 0:
        reason_groups = {}
        for idx, (_, row) in enumerate(current_alerts.iterrows()):
            reason = row.get('告警原因', '')
            if pd.isna(reason) or not str(reason).strip():
                continue
            reason = str(reason).strip()
            if reason not in reason_groups:
                reason_groups[reason] = []
            reason_groups[reason].append(idx + 1)

        for reason, alert_nums in reason_groups.items():
            nums_str = '、'.join([get_alert_circle_number(n) for n in alert_nums])
            text = f"アラート{nums_str}：{reason}"
            ws.cell(row=current_row, column=1).value = text
            current_row += 1

    # 空一行
    current_row += 1

    # 3、リソース利用状況
    ws.cell(row=current_row, column=1).value = '3、リソース利用状況'
    current_row += 1
    ws.cell(row=current_row, column=1).value = '3-1リソース使用率一覧'
    current_row += 1

    # 资源表头
    ws.cell(row=current_row, column=1).value = '環境'
    ws.cell(row=current_row, column=2).value = 'ホスト名'
    ws.cell(row=current_row, column=4).value = 'CPU使用率'
    ws.cell(row=current_row, column=8).value = 'Memory使用率'
    ws.cell(row=current_row, column=12).value = 'Disk使用率'
    current_row += 1

    # 资源数据行（带环境分组）
    if not cmd_data.empty:
        current_row = _fill_resource_table(ws, cmd_data, current_row, host_mapping)

    # BarChart
    _create_alert_chart(wb, ws)

    logger.info(f"【総評】填充完成（动态布局，数据区从R{clear_from}起）")


def _clear_dynamic_area(ws, from_row):
    """清空指定行以下的所有内容和合并单元格"""
    # 取消合并单元格
    merged_to_remove = []
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row >= from_row:
            merged_to_remove.append(merged_range)
    for mr in merged_to_remove:
        ws.unmerge_cells(str(mr))

    # 清空内容
    max_row = ws.max_row
    max_col = ws.max_column
    for r in range(from_row, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).value = None


def _fill_resource_table(ws, cmd_data, start_row, host_mapping):
    """填充资源使用率表（带环境分组和显示名），返回下一可用行"""
    current_row = start_row
    current_env = None

    for _, row in cmd_data.iterrows():
        instance_id = str(row.get('InstanceId', row.get('InstanceName', '')))
        instance_name = str(row.get('InstanceName', ''))

        # 查找显示名和环境
        host_info = host_mapping.get(instance_id, host_mapping.get(instance_name, {}))
        display_name = host_info.get('display_name', instance_name)
        env = host_info.get('env', '')

        # 环境分组标题行
        if env and env != current_env:
            current_env = env
            ws.cell(row=current_row, column=1).value = env
            current_row += 1

        # 数据行
        ws.cell(row=current_row, column=2).value = display_name
        ws.cell(row=current_row, column=4).value = row.get('CPU(核)', '')
        ws.cell(row=current_row, column=5).value = row.get('CPU最小使用率(%)', '')
        ws.cell(row=current_row, column=6).value = row.get('CPU最大使用率(%)', '')
        ws.cell(row=current_row, column=7).value = row.get('CPU平均使用率(%)', '')
        ws.cell(row=current_row, column=8).value = row.get('Memory(GB)', '')
        ws.cell(row=current_row, column=9).value = row.get('内存最小使用率(%)', '')
        ws.cell(row=current_row, column=10).value = row.get('内存最大使用率(%)', '')
        ws.cell(row=current_row, column=11).value = row.get('内存平均使用率(%)', '')
        ws.cell(row=current_row, column=12).value = row.get('磁盘大小(GB)', '')
        ws.cell(row=current_row, column=13).value = row.get('磁盘最大使用率（%)', '')
        ws.cell(row=current_row, column=15).value = row.get('磁盘剩余大小(GB)', '')
        ws.cell(row=current_row, column=16).value = row.get('磁盘数据最大使用率(%)', '')
        current_row += 1

    logger.info(f"【総評】资源表填充完成，共{len(cmd_data)}行（含环境分组）")
    return current_row


def _create_alert_chart(wb, summary_ws):
    """在総評 sheet 创建告警件数 BarChart，引用推移 sheet"""
    if SHEET_TREND not in wb.sheetnames:
        return
    trend_ws = wb[SHEET_TREND]

    # 找到推移sheet中有数据的最后一行
    max_data_row = TREND_DATA_START_ROW
    for r in range(TREND_DATA_START_ROW, trend_ws.max_row + 1):
        if trend_ws.cell(row=r, column=2).value:
            max_data_row = r
        else:
            break

    if max_data_row < TREND_DATA_START_ROW:
        return

    chart = BarChart()
    chart.type = "col"
    chart.title = "全サーバアラート対応件数"
    chart.style = 10
    chart.width = 20
    chart.height = 16

    # C=前月, D=今月
    data_ref = Reference(trend_ws, min_col=3, min_row=TREND_HEADER_ROW,
                         max_col=4, max_row=max_data_row)
    cats_ref = Reference(trend_ws, min_col=2, min_row=TREND_DATA_START_ROW,
                         max_row=max_data_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.shape = 4

    # 坐标轴可见
    chart.y_axis.tickLblPos = 'nextTo'

    # 将 X 轴的原标签隐藏，因为要用底部的数据表(Data Table)代替
    chart.x_axis.tickLblPos = 'none'

    # 图例隐藏
    chart.legend = None

    # 开启底部数据表 DataTable
    dt = DataTable()
    dt.showHorzBorder = True  # 显示水平网格
    dt.showVertBorder = True  # 显示垂直网格
    dt.showOutline = True  # 显示外边框
    dt.showKeys = True  # 显示色块图例
    chart.plotArea.dTable = dt

    summary_ws.add_chart(chart, LAYOUT['chart_anchor'])
    # 强制设置图表尺寸（直接构建 OneCellAnchor）
    _set_chart_size(summary_ws._charts[-1], LAYOUT['chart_anchor'], 20, 16)
    logger.info(f"【総評】BarChart创建完成")


def populate_trend_sheet(wb, host_list, current_alerts, prev_alerts):
    """填充推移 sheet (C=前月, D=今月)"""
    if SHEET_TREND not in wb.sheetnames:
        return
    ws = wb[SHEET_TREND]

    # 先取消数据区域的合并单元格
    merged_to_remove = []
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row >= TREND_DATA_START_ROW:
            merged_to_remove.append(merged_range)
    for mr in merged_to_remove:
        ws.unmerge_cells(str(mr))

    # 清除旧数据
    for r in range(TREND_DATA_START_ROW, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).value = None

    # 表头: B=対象, C=前月, D=今月
    ws.cell(row=TREND_HEADER_ROW, column=2).value = '対象'
    ws.cell(row=TREND_HEADER_ROW, column=3).value = '前月'
    ws.cell(row=TREND_HEADER_ROW, column=4).value = '今月'

    for idx, obj_name in enumerate(host_list):
        row_num = TREND_DATA_START_ROW + idx
        ws.cell(row=row_num, column=2).value = obj_name

        # 前月告警数 (C列)
        prev_count = 0
        if not prev_alerts.empty:
            prev_count = len(prev_alerts[
                                 prev_alerts['告警目标'].astype(str).str.lower() == obj_name.lower()
                                 ])
        ws.cell(row=row_num, column=3).value = prev_count

        # 今月告警数 (D列)
        current_count = 0
        if not current_alerts.empty:
            current_count = len(current_alerts[
                                    current_alerts['告警目标'].astype(str).str.lower() == obj_name.lower()
                                    ])
        ws.cell(row=row_num, column=4).value = current_count

    # 合計行
    total_row = TREND_DATA_START_ROW + len(host_list) + 1  # 空一行
    ws.cell(row=total_row, column=2).value = '合計'
    first_data = TREND_DATA_START_ROW
    last_data = TREND_DATA_START_ROW + len(host_list) - 1
    ws.cell(row=total_row, column=3).value = f'=SUM(C{first_data}:C{last_data})'
    ws.cell(row=total_row, column=4).value = f'=SUM(D{first_data}:D{last_data})'

    logger.info(f"【推移】填充完成，共{len(host_list)}个对象")


def populate_resource_sheet(wb, sheet_name, metric_type, host_list, data_dir,
                            account, host_mapping):
    """填充リソース sheet (CPU/Memory/Disk)，使用显示名"""
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]

    metric_label = {'cpu': 'CPU', 'memory': 'Memory', 'disk': 'Disk'}[metric_type]
    current_row = RESOURCE_TITLE_START_ROW
    data_col = RESOURCE_DATA_START_COL

    for host in host_list:
        # 获取显示名
        host_info = host_mapping.get(host, {})
        display_name = host_info.get('display_name', host)

        # 标题
        ws.cell(row=current_row, column=2).value = f"■{display_name}({metric_label}使用率)"

        # 加载监控数据
        monitor_df = load_monitoring_data(data_dir, account, host, metric_type)

        if not monitor_df.empty and 'datetime' in monitor_df.columns:
            header_row = current_row
            ws.cell(row=header_row, column=data_col).value = 'datetime'
            ws.cell(row=header_row, column=data_col + 1).value = 'average'

            if metric_type == 'disk' and '盘符' in monitor_df.columns:
                disk_groups = monitor_df.groupby('盘符')['average'].mean()
                if not disk_groups.empty:
                    monitor_df = monitor_df[monitor_df['盘符'] == disk_groups.idxmax()]

            num_points = len(monitor_df)
            for pt_idx, (_, pt) in enumerate(monitor_df.iterrows()):
                r = header_row + 1 + pt_idx
                ws.cell(row=r, column=data_col).value = pt['datetime'].strftime('%m/%d %H:%M')
                ws.cell(row=r, column=data_col + 1).value = pt.get('average', 0)

            chart = LineChart()
            chart.title = f"{display_name} ({metric_label}使用率)"
            chart.style = 10
            chart.width = 22
            chart.height = 8
            chart.y_axis.title = '%'
            chart.x_axis.title = '日時'
            chart.y_axis.tickLblPos = 'nextTo'
            chart.x_axis.tickLblPos = 'low'
            chart.legend = None  # 单系列不需要图例

            # 只取 average 一列
            data_ref = Reference(ws, min_col=data_col + 1, min_row=header_row,
                                 max_col=data_col + 1, max_row=header_row + num_points)
            cats_ref = Reference(ws, min_col=data_col, min_row=header_row + 1,
                                 max_row=header_row + num_points)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            ws.add_chart(chart, f"B{current_row + 1}")
            # 强制设置图表尺寸
            _set_chart_size(ws._charts[-1], f"B{current_row + 1}", 22, 8)
        else:
            logger.warning(f"  [{sheet_name}] {display_name}: 无监控数据")

        current_row += RESOURCE_INTERVAL_ROWS

    logger.info(f"【{sheet_name}】填充完成，共{len(host_list)}台主机")


# --- V. 主逻辑 ---
def build_project_report(data, target_year, target_month, today):
    """为 TSIINFRA 项目生成运维月报"""
    project_key = PROJECT_KEY

    # 名称映射
    names_df = data.get('names', pd.DataFrame())
    name_row = names_df[names_df['Project Key'] == project_key] if not names_df.empty else pd.DataFrame()
    if not name_row.empty:
        company_name = str(name_row.iloc[0]['公司名'])
        system_name = str(name_row.iloc[0]['系统名'])
    else:
        logger.warning(f"项目 {project_key} 无名称映射，跳过")
        return None

    logger.info(f"\n{'=' * 60}")
    logger.info(f"开始处理项目: {project_key} ({company_name})")
    logger.info(f"{'=' * 60}")

    # 主机映射
    host_mapping = get_host_mapping(data.get('hosts', pd.DataFrame()), project_key)

    # 账号和标签
    account_df = data['op_account']
    proj_accounts = account_df[account_df['Project Key'] == project_key]
    if proj_accounts.empty:
        logger.warning(f"项目 {project_key} 无对应账号，跳过")
        return None

    account_row = proj_accounts.iloc[0]
    account = account_row['资源所属账号']
    tag_filter = parse_account_tag(account_row.get('标签', ''))
    logger.info(f"账号: {account}, 标签过滤: {tag_filter}")

    # ECS实例
    instances_df = get_project_instances(account, tag_filter, data['ecs'])
    if instances_df.empty:
        logger.warning(f"项目 {project_key} 无匹配ECS实例，跳过")
        return None

    # 去重
    instances_df = instances_df.drop_duplicates(subset=['InstanceId'], keep='first')
    host_list = instances_df['InstanceName'].tolist()
    logger.info(f"匹配到 {len(host_list)} 台ECS实例(去重后)")

    # 资源使用率汇总
    cmd_df = data['cmd_summary']
    cmd_proj = cmd_df[cmd_df['资源所属账号'] == account].copy()
    cmd_proj = cmd_proj[cmd_proj['InstanceName'].isin(host_list)]

    if not cmd_proj.empty:
        ecs_info = instances_df[['InstanceName', 'InstanceId', 'CPU(核)', 'Memory(MB)']].copy()
        ecs_info['Memory(GB)'] = (ecs_info['Memory(MB)'] / 1024).round(0).astype(int)
        cmd_proj = cmd_proj.merge(
            ecs_info[['InstanceName', 'InstanceId', 'CPU(核)', 'Memory(GB)']],
            on='InstanceName', how='left'
        )
        cmd_proj = cmd_proj.drop_duplicates(subset=['InstanceName'], keep='first')

        # 按环境映射排序
        if host_mapping:
            env_order = {}
            for iid, info in host_mapping.items():
                env_order[info['display_name']] = list(host_mapping.keys()).index(iid)
            cmd_proj['sort_key'] = cmd_proj['InstanceName'].map(
                lambda x: env_order.get(x, 999)
            )
            cmd_proj = cmd_proj.sort_values('sort_key').drop(columns=['sort_key'])

    # 告警数据
    current_alerts, prev_alerts = get_alerts_for_project(
        data['alert'], project_key, target_year, target_month
    )
    logger.info(f"告警(过滤后): 今月{len(current_alerts)}件, 前月{len(prev_alerts)}件")

    # 加载模板
    template_path = data['template_path']
    logger.info(f"使用模板: {os.path.basename(template_path)}")
    wb = load_workbook(template_path)

    # 填充各 Sheet
    populate_cover(wb, company_name, system_name, target_year, target_month, today)

    # 推移 sheet（使用显示名）
    trend_display_list = []
    for _, inst_row in instances_df.iterrows():
        iid = str(inst_row['InstanceId'])
        iname = str(inst_row['InstanceName'])
        info = host_mapping.get(iid, host_mapping.get(iname, {}))
        trend_display_list.append(info.get('display_name', iname))

    populate_trend_sheet(wb, trend_display_list, current_alerts, prev_alerts)

    populate_summary(wb, current_alerts, prev_alerts, cmd_proj,
                     target_year, target_month, host_mapping)

    # リソース sheets
    populate_resource_sheet(wb, SHEET_CPU, 'cpu', host_list, data['data_dir'],
                            account, host_mapping)
    populate_resource_sheet(wb, SHEET_MEMORY, 'memory', host_list, data['data_dir'],
                            account, host_mapping)
    populate_resource_sheet(wb, SHEET_DISK, 'disk', host_list, data['data_dir'],
                            account, host_mapping)

    # 保存
    output_dir = data['output_dir']
    project_folder = os.path.join(output_dir, project_key)
    os.makedirs(project_folder, exist_ok=True)

    year_month_str = f"{target_year}{target_month:02d}"
    output_filename = f"【{company_name}様向け】インフラ月次報告書({year_month_str}).xlsx"
    output_path = os.path.join(project_folder, output_filename)

    wb.save(output_path)
    wb.close()

    # 修正图表尺寸（openpyxl 保存时不正确写入 anchor ext）
    # drawing1=総評 BarChart, drawing2-4=リソース LineCharts
    _fix_chart_sizes_in_xlsx(output_path, {
        1: (20, 16),  # 総評 BarChart: 20cm x 16cm
        2: (22, 8),  # リソース (CPU): 22cm x 8cm
        3: (22, 8),  # リソース (Memory)
        4: (22, 8),  # リソース (Disk)
    })

    logger.info(f"报表生成成功: {output_path}")
    return output_path


def main():
    logger.info("=" * 80)
    logger.info(f"开始执行 TSIINFRA 运维月报生成: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    try:
        today = datetime.today()
        target_date = today.replace(day=1) - timedelta(days=1)
        target_year = target_date.year
        target_month = target_date.month
        logger.info(f"目标月份: {target_year}年{target_month}月")

        script_dir = os.path.dirname(os.path.abspath(__file__))
        data = load_all_data(script_dir)

        result = build_project_report(data, target_year, target_month, today)
        if result:
            logger.info(f"\n{'=' * 80}")
            logger.info(f"TSIINFRA 报表生成成功: {result}")
        else:
            logger.error("TSIINFRA 报表生成失败")

        logger.info(f"结束时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    except Exception as e:
        logger.error(f"程序执行失败: {str(e)}", exc_info=True)


if __name__ == "__main__":
    main()