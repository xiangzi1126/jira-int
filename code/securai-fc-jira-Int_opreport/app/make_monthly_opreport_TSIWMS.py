import os
import re
import glob
import math
import zipfile
import shutil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.axis import DateAxis
from openpyxl.chart.series import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText
from openpyxl.chart.plotarea import DataTable
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.fill import ColorChoice
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, RichTextProperties
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import cm_to_EMU
from openpyxl.styles import Font
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import logging

# 総評シート挿入データのフォント (MS PGothic 10pt) - UNICYCEN/TSIINFRAと同一
DATA_FONT = Font(name='MS PGothic', size=10)

# --- I. 常量定义 ---
PROJECT_KEY = 'TSIWMS'

SHEET_COVER = '表紙'
SHEET_SUMMARY = '総評1-8'
SHEET_CPU = 'リソース (CPU)'
SHEET_MEMORY = 'リソース (Memory)'
SHEET_DISK = 'リソース (Disk )'
SHEET_RDS = 'RDS'
SHEET_TREND = '推移'

# 表紙 常量
COVER_COMPANY_ROW = 3
COVER_SYSTEM_ROW = 8
COVER_DATE_ROW = 29

# 推移 常量
TREND_HEADER_ROW = 2
TREND_DATA_START_ROW = 3

# リソース 常量
RESOURCE_TITLE_START_ROW = 3
RESOURCE_INTERVAL_ROWS = 18
RESOURCE_DATA_START_COL = 20  # T列

# WMS リソース図表尺寸 (模板 7*20.7cm 注記: 高7cm x 幅20.7cm。模板タイトル間隔が非均一16/18/20行)
RESOURCE_CHART_W = 20.7
RESOURCE_CHART_H = 7

# 総評 布局配置 (WMS 模板固定位置。TSIINFRA方式: 挿入行でoffset累積)
# WMS 模板与 TSIINFRA 行号不同: ■検知内容 data=R49(header R48), ■事象纏め data=R52(header R51), 【CPU】=R92
LAYOUT = {
    'greeting_row': 2,           # R2 月份挨拶
    'overview_row': 5,           # R5 1、総評 概要 (模板既定文="全体的に安定稼働しております。")
    'alert_count_row': 8,        # R8 2-1アラート件数
    'detection_row': 49,         # R49 ■検知内容 数据行 (header R48, 既定文="特にございません。")
    'incident_row': 52,          # R52 ■事象纏め 数据行 (header R51, 既定文="特にございません。")
    'resource_clear_from': 54,   # R54 3、リソース利用状況 (定位填入模板, 不清空重建)
    'comments_row': 91,          # R91 【CPU】(R92)前空行
}


# --- II. 辅助函数 ---
def init_logger():
    log_dir = os.path.normpath(os.path.join(os.path.dirname(__file__), '../../jira/log'))
    os.makedirs(log_dir, exist_ok=True)
    log_file = os.path.join(log_dir, 'monthly_opreport_TSIWMS.log')
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler()
        ]
    )
    return logging.getLogger(__name__)


logger = init_logger()


def safe_read_csv(file_path):
    """多编码尝试读取CSV"""
    for encoding in ['utf-8-sig', 'utf-8', 'gbk', 'shift-jis', 'latin1']:
        try:
            return pd.read_csv(file_path, encoding=encoding)
        except (UnicodeDecodeError, UnicodeError):
            continue
        except Exception as e:
            logger.warning(f"读取文件 {file_path} 失败: {e}")
            continue
    raise IOError(f"无法读取文件: {file_path}")


def parse_account_tag(tag_str):
    """将 'key:TSI value:infra' 转换为 'TSI:infra'"""
    if not tag_str or pd.isna(tag_str) or str(tag_str).strip() in ('', 'nan'):
        return None
    tag_str = str(tag_str).strip()
    key_match = re.search(r'key:(\S+)', tag_str)
    val_match = re.search(r'value:(\S+)', tag_str)
    if key_match and val_match:
        return f"{key_match.group(1)}:{val_match.group(1)}"
    return tag_str


def get_alert_circle_number(n):
    """数字转圆圈数字 ①②③..."""
    circles = '①②③④⑤⑥⑦⑧⑨⑩⑪⑫⑬⑭⑮⑯⑰⑱⑲⑳'
    if 1 <= n <= 20:
        return circles[n - 1]
    return f'({n})'


def _set_chart_size(chart_obj, anchor_str, width_cm, height_cm):
    """设置图表尺寸（内存中设置，保存后用 _fix_chart_sizes_in_xlsx 修正 XML）"""
    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
    col_letter, row = coordinate_from_string(anchor_str)
    col = column_index_from_string(col_letter)
    marker = AnchorMarker(col=col - 1, colOff=0, row=row - 1, rowOff=0)
    size = XDRPositiveSize2D(cx=cm_to_EMU(width_cm), cy=cm_to_EMU(height_cm))
    chart_obj.anchor = OneCellAnchor(_from=marker, ext=size)


def _fix_chart_sizes_in_xlsx(xlsx_path, size_map):
    """保存后修正 xlsx 内 drawing XML 的图表尺寸
    size_map: {drawing序号(1-based): (width_cm, height_cm)}
    对该 drawing 内的所有 ext 统一替换
    """
    tmp_path = xlsx_path + '.tmp'
    pat = re.compile(r'(<ext cx=")([0-9]+)(" cy=")([0-9]+)(")')

    with zipfile.ZipFile(xlsx_path, 'r') as zin:
        with zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                m = re.match(r'xl/drawings/drawing(\d+)\.xml', item.filename)
                if m:
                    drawing_idx = int(m.group(1))
                    if drawing_idx in size_map:
                        w_cm, h_cm = size_map[drawing_idx]
                        new_cx = str(cm_to_EMU(w_cm))
                        new_cy = str(cm_to_EMU(h_cm))
                        text = data.decode('utf-8')
                        text = pat.sub(
                            lambda g: g.group(1) + new_cx + g.group(3) + new_cy + g.group(5),
                            text
                        )
                        data = text.encode('utf-8')
                zout.writestr(item, data)

    # Windows 安全替换
    os.remove(xlsx_path)
    shutil.move(tmp_path, xlsx_path)
    logger.info(f"图表尺寸修正完成: {size_map}")


def safe_insert_rows(ws, idx, n):
    """insert_rows 安全版。
    openpyxl 3.1.x 的 insert_rows 会平移单元格的值/样式, 但合并单元格只新建平移后的、
    不删除旧的 -> 残留重叠合并(Q49:R49 等"错乱")。
    本函数: 插入前 unmerge 插入点及下方(含跨越插入点)的合并, 插入后按平移坐标重新 merge。"""
    if n <= 0:
        return
    # 收集受影响合并的原始坐标 (max_row >= idx: 整体在下方 或 跨越插入点)
    affected = [(mr.min_col, mr.min_row, mr.max_col, mr.max_row)
                for mr in list(ws.merged_cells.ranges) if mr.max_row >= idx]
    # 先 unmerge, 避免 insert_rows 留下旧合并
    for mr in list(ws.merged_cells.ranges):
        if mr.max_row >= idx:
            ws.unmerge_cells(str(mr))
    # 插入行 (openpyxl 会平移值/样式)
    ws.insert_rows(idx, n)
    # 按平移后坐标重新合并
    for c1, r1, c2, r2 in affected:
        new_r1 = r1 + n if r1 >= idx else r1
        new_r2 = r2 + n
        ws.merge_cells(start_row=new_r1, start_column=c1, end_row=new_r2, end_column=c2)


# --- III. 数据加载 ---
def load_all_data(script_dir):
    """加载所有数据源"""
    def get_path(rel_path):
        return os.path.normpath(os.path.join(script_dir, rel_path))

    data_dir = get_path("../../jira/data")
    output_dir = get_path("../../jira/monthly_report")
    template_path = get_path(
        "../../jira/monthly_report/运维月报模板_【株式会社TSI様向け】WMS月次報告書.xlsx"
    )
    os.makedirs(output_dir, exist_ok=True)

    # 数据文件
    op_details_csv = os.path.join(data_dir, 'op_resale_business_details.csv')
    op_account_csv = os.path.join(data_dir, 'jira_get_op_account.csv')
    ecs_csv = os.path.join(data_dir, 'aliyun_ecs_op.csv')
    cmd_summary_csv = os.path.join(data_dir, 'aliyun_op_cmd_summary.csv')
    alert_csv = os.path.join(data_dir, 'jira_get_alert.csv')

    for path in [op_details_csv, op_account_csv, ecs_csv, cmd_summary_csv, alert_csv]:
        if not os.path.exists(path):
            raise FileNotFoundError(f"关键文件不存在: {path}")

    if not os.path.exists(template_path):
        raise FileNotFoundError(f"模板文件不存在: {template_path}")

    op_details_df = safe_read_csv(op_details_csv)
    op_account_df = safe_read_csv(op_account_csv)
    ecs_df = safe_read_csv(ecs_csv)
    cmd_summary_df = safe_read_csv(cmd_summary_csv)
    alert_df = safe_read_csv(alert_csv)

    # RDS 数据 (op_rds: 实例盘点 CPU核/memory(GB)/disk(GB); rds_cmd: CPU/Mem/Disk 使用率汇总)
    rds_op_csv = os.path.join(data_dir, 'aliyun_rds_op.csv')
    rds_cmd_csv = os.path.join(data_dir, 'aliyun_rds_cmd_summary.csv')
    rds_op_df = safe_read_csv(rds_op_csv) if os.path.exists(rds_op_csv) else pd.DataFrame()
    rds_cmd_df = safe_read_csv(rds_cmd_csv) if os.path.exists(rds_cmd_csv) else pd.DataFrame()

    # 配置文件
    names_csv = get_path("../../jira/config/opreport_names.csv")
    hosts_csv = get_path("../../jira/config/opreport_hosts.csv")
    names_df = safe_read_csv(names_csv) if os.path.exists(names_csv) else pd.DataFrame()
    hosts_df = safe_read_csv(hosts_csv) if os.path.exists(hosts_csv) else pd.DataFrame()

    # SVFS ディスク最大拡張容量 + 数据盘顺序（人工月报记录的值，cmd_summary 无此字段）
    # WMS 无 SVFS 主机, 此 map 通常为空 (保留与 TSIINFRA 一致的结构)
    svfs_expand_csv = get_path("../../jira/config/opreport_svfs_disk_expand.csv")
    svfs_expand_map = {}
    if os.path.exists(svfs_expand_csv):
        sdf = safe_read_csv(svfs_expand_csv)
        for _, srow in sdf.iterrows():
            host = str(srow.get('InstanceName', '')).lower().strip()
            panfu = str(srow.get('盘符', '')).strip()
            drive = panfu[0].upper() if panfu else ''
            val = srow.get('ディスク最大拡張容量(GB)', '-')
            svfs_expand_map.setdefault(host, []).append((drive, val))

    logger.info(f"数据加载完成: 项目{len(op_details_df)}个, 账号{len(op_account_df)}个, "
                f"ECS{len(ecs_df)}台, 资源汇总{len(cmd_summary_df)}条, 告警{len(alert_df)}条, "
                f"RDS{len(rds_op_df)}台/{len(rds_cmd_df)}条汇总")
    if not names_df.empty:
        svfs_drive_cnt = sum(len(v) for v in svfs_expand_map.values())
        logger.info(f"名称映射: {len(names_df)}条, 主机映射: {len(hosts_df)}条, "
                    f"SVFS拡張容量: {svfs_drive_cnt}条")

    return {
        'output_dir': output_dir,
        'template_path': template_path,
        'data_dir': data_dir,
        'op_details': op_details_df,
        'op_account': op_account_df,
        'ecs': ecs_df,
        'cmd_summary': cmd_summary_df,
        'alert': alert_df,
        'names': names_df,
        'hosts': hosts_df,
        'svfs_expand': svfs_expand_map,
        'rds_op': rds_op_df,
        'rds_cmd': rds_cmd_df,
    }


def get_project_instances(account, tag_filter, ecs_df):
    """根据账号和标签过滤ECS实例"""
    account_df = ecs_df[ecs_df['资源所属账号'] == account]
    if tag_filter:
        tag_mask = account_df['标签'].astype(str).str.contains(re.escape(tag_filter), na=False)
        account_df = account_df[tag_mask]
    return account_df


def get_alerts_for_project(alert_df, project_key, target_year, target_month):
    """获取项目的今月和前月告警（排除 备注='不记录月报'）"""
    proj_alerts = alert_df[alert_df['Project Key'] == project_key].copy()
    if proj_alerts.empty:
        return pd.DataFrame(), pd.DataFrame()

    # 过滤：排除 备注="不记录月报"
    if '备注' in proj_alerts.columns:
        proj_alerts = proj_alerts[
            proj_alerts['备注'].astype(str).str.strip() != '不记录月报'
        ]

    if proj_alerts.empty:
        return pd.DataFrame(), pd.DataFrame()

    # 解析告警检测日时（转换为JST后去除时区）
    proj_alerts['告警检测日时'] = pd.to_datetime(
        proj_alerts['告警检测日时'], errors='coerce', utc=True
    ).dt.tz_convert('Asia/Tokyo').dt.tz_localize(None)
    proj_alerts = proj_alerts.dropna(subset=['告警检测日时'])

    # 今月/前月范围
    current_start = datetime(target_year, target_month, 1)
    current_end = current_start + relativedelta(months=1) - timedelta(seconds=1)
    prev_start = current_start - relativedelta(months=1)
    prev_end = current_start - timedelta(seconds=1)

    current_alerts = proj_alerts[
        (proj_alerts['告警检测日时'] >= current_start) &
        (proj_alerts['告警检测日时'] <= current_end)
    ].sort_values('告警检测日时')
    prev_alerts = proj_alerts[
        (proj_alerts['告警检测日时'] >= prev_start) &
        (proj_alerts['告警检测日时'] <= prev_end)
    ]
    return current_alerts, prev_alerts


def load_monitoring_data(data_dir, account, instance_name, metric_type):
    """加载单个主机的监控时序数据"""
    monitor_dir = os.path.join(data_dir, f'监控数据_{account}')
    if not os.path.exists(monitor_dir):
        return pd.DataFrame()

    pattern = os.path.join(monitor_dir, f'{metric_type}_*.csv')
    files = glob.glob(pattern)

    target_file = None
    for f in files:
        basename = os.path.basename(f)
        name_part = basename.replace(f'{metric_type}_', '').replace('.csv', '')
        if name_part.lower() == instance_name.lower():
            target_file = f
            break

    if not target_file:
        return pd.DataFrame()

    try:
        df = safe_read_csv(target_file)
        if 'timestamp' in df.columns:
            df['datetime'] = pd.to_datetime(df['timestamp'], unit='ms')
            df = df.sort_values('datetime')
        return df
    except Exception as e:
        logger.warning(f"读取监控数据失败 {target_file}: {e}")
        return pd.DataFrame()


def load_rds_monitoring_data(data_dir, account, instance_name, metric_type):
    """加载单个 RDS 实例的监控时序数据 (文件名 rds_{metric}_{name}.csv, 带 rds_ 前缀)"""
    monitor_dir = os.path.join(data_dir, f'监控数据_{account}')
    if not os.path.exists(monitor_dir):
        return pd.DataFrame()

    pattern = os.path.join(monitor_dir, f'rds_{metric_type}_*.csv')
    files = glob.glob(pattern)

    target_file = None
    for f in files:
        basename = os.path.basename(f)
        name_part = basename.replace(f'rds_{metric_type}_', '').replace('.csv', '')
        if name_part.lower() == instance_name.lower():
            target_file = f
            break

    if not target_file:
        return pd.DataFrame()

    try:
        df = safe_read_csv(target_file)
        if 'timestamp' in df.columns:
            df['datetime'] = pd.to_datetime(df['timestamp'], unit='ms')
            df = df.sort_values('datetime')
        return df
    except Exception as e:
        logger.warning(f"读取RDS监控数据失败 {target_file}: {e}")
        return pd.DataFrame()


def get_host_mapping(hosts_df, project_key):
    """获取项目的主机映射表（显示名、环境）"""
    if hosts_df.empty:
        return {}
    proj_hosts = hosts_df[hosts_df['Project Key'] == project_key]
    mapping = {}
    for _, row in proj_hosts.iterrows():
        mapping[str(row['InstanceId'])] = {
            'display_name': row['显示名'],
            'env': row['环境'],
        }
    return mapping


# --- IV. Sheet 填充函数 ---
def populate_cover(wb, company_name, system_name, target_year, target_month, today):
    """填充表紙"""
    if SHEET_COVER not in wb.sheetnames:
        return
    ws = wb[SHEET_COVER]
    ws.cell(row=COVER_COMPANY_ROW, column=1).value = f"{company_name}\u3000御中"
    ws.cell(row=COVER_SYSTEM_ROW, column=1).value = (
        f"{system_name}\n{target_year}年{target_month}月\u3000月次報告"
    )
    ws.cell(row=COVER_DATE_ROW, column=1).value = (
        f"\n作成日：{today.year}/{today.month}/{today.day}\nセキュライ大連"
    )
    logger.info(f"【表紙】填充完成: {company_name}, {target_year}年{target_month}月")


def populate_summary(wb, current_alerts, prev_alerts, cmd_data,
                     target_year, target_month, host_mapping, today,
                     svfs_expand_map, c11_map, rds_op_df, rds_cmd_df,
                     account, tag_filter):
    """填充総評1-8（动态布局：插入行+定位填入模板。WMS 无 課金一覧, 課金填充为 no-op）"""
    if SHEET_SUMMARY not in wb.sheetnames:
        return
    ws = wb[SHEET_SUMMARY]

    def set_cell(row, col, value):
        """セルに値を書き込み + フォント(MS PGothic 10pt) - UNICYCEN/TSIINFRAと同一"""
        c = ws.cell(row=row, column=col)
        c.value = value
        c.font = DATA_FONT

    # R2: 月份挨拶 (R1挨拶文は模板固定)
    set_cell(LAYOUT['greeting_row'], 1,
             f"{target_year}年{target_month}月のシステム稼働状況について、下記の通り、ご報告申し上げます。")

    alert_count = len(current_alerts)
    offset = 0  # 行挿入による累積オフセット (UNICYCEN方式)

    # === 1、総評 概要 (R5) ===
    # 0件: テンプレ既定文 "全体的に安定稼働しております。" を維持
    # N件: R5クリア -> R6に(N-1)行挿入 -> 各告警の备注を "M/D 、<备注>" で出力
    if alert_count > 0:
        set_cell(LAYOUT['overview_row'], 1, None)
        if alert_count > 1:
            insert_at = LAYOUT['overview_row'] + 1   # R6
            insert_n = alert_count - 1
            safe_insert_rows(ws, insert_at, insert_n)
            offset += insert_n
            logger.info(f"【総評】総評概要に{insert_n}行挿入")
        for idx, (_, row) in enumerate(current_alerts.iterrows()):
            r = LAYOUT['overview_row'] + idx
            detect_time = row['告警检测日时']
            remark = row.get('备注', '')
            if pd.isna(remark) or not str(remark).strip():
                remark = f"{row.get('告警目标', '')}サーバのアラートを検知しました。"
            date_str = f"{detect_time.month}/{detect_time.day}"
            set_cell(r, 1, f"{date_str} 、{str(remark).strip()}")

    # 2-1アラート件数 (R8+offset。"2、アラート対応状況"は模板固定)
    set_cell(LAYOUT['alert_count_row'] + offset, 1, f"2-1アラート件数：{alert_count}件")

    # === ■検知内容 (R48 header固定, R49 data) ===
    detection_data_start = LAYOUT['detection_row'] + offset
    if alert_count <= 0:
        set_cell(detection_data_start, 1, "特にございません。")
    else:
        set_cell(detection_data_start, 1, None)
        if alert_count > 1:
            insert_at = detection_data_start + 1
            insert_n = alert_count - 1
            safe_insert_rows(ws, insert_at, insert_n)
            offset += insert_n
            logger.info(f"【総評】検知内容に{insert_n}行挿入")
        for idx, (_, row) in enumerate(current_alerts.iterrows()):
            r = detection_data_start + idx
            detect_time = row['告警检测日时']
            alert_desc = str(row.get('告警简述', '') or '').strip()
            date_str = (f"{detect_time.year}-{detect_time.month}-{detect_time.day} "
                        f"{detect_time.hour}:{detect_time.minute:02d}")
            circle_num = get_alert_circle_number(idx + 1)
            text = f"アラート{circle_num}：{date_str} {alert_desc}".strip()
            set_cell(r, 1, text)

    # === ■事象纏め (R51 header固定, R52 data) ===
    incident_data_row = LAYOUT['incident_row'] + offset
    reason_groups = {}
    if alert_count > 0:
        for idx, (_, row) in enumerate(current_alerts.iterrows()):
            reason = row.get('告警原因', '')
            if pd.isna(reason) or not str(reason).strip():
                continue
            reason = str(reason).strip()
            if reason not in reason_groups:
                reason_groups[reason] = []
            reason_groups[reason].append(idx + 1)

    if alert_count <= 0 or not reason_groups:
        set_cell(incident_data_row, 1, "特にございません。")
    else:
        set_cell(incident_data_row, 1, None)
        reasons = list(reason_groups.items())
        if len(reasons) > 1:
            insert_at = incident_data_row + 1
            insert_n = len(reasons) - 1
            safe_insert_rows(ws, insert_at, insert_n)
            offset += insert_n
            logger.info(f"【総評】事象纏めに{insert_n}行挿入")
        for ridx, (reason, alert_nums) in enumerate(reasons):
            r = incident_data_row + ridx
            nums_str = '、'.join([get_alert_circle_number(n) for n in alert_nums])
            set_cell(r, 1, f"アラート{nums_str}：{reason}")

    # === 资源表: 定位填入模板(保留合并单元格), 不清空重建 ===
    # 模板的 3、リソース/3-1/表头/合并/sections 都保留, 只往主机名行填数据
    # WMS 资源表列: L=系统盘大小, M=系统盘max, N=数据盘大小, O=数据盘max (无 M:N/P:Q 合并)
    if not cmd_data.empty:
        _fill_resource_table_inplace(ws, cmd_data, host_mapping)
        _fill_svfs_table_inplace(ws, cmd_data, svfs_expand_map)

    # RDS 子表 (R80-81): 独立于 ECS cmd_data, 用 rds_op + rds_cmd 填, 模仿 ECS 定位填入
    _fill_rds_table_inplace(ws, rds_op_df, rds_cmd_df, account, tag_filter)

    # === section 6 M月/N月 占位符置換 (6-1 M月の作業状況 -> 当月, 6-2 N月の作業予定 -> 次月) ===
    _replace_section_months(ws, target_year, target_month)

    # 課金 更新日(C7) = 作成月第一天, C11=最新有効期限
    # WMS 総評1-8 无 課金一覧 section -> _fill_kaiin_renewal_date 为 no-op (保留调用, 与 TSIINFRA 一致)
    renewal_date = datetime(today.year, today.month, 1)
    _fill_kaiin_renewal_date(ws, renewal_date, c11_map)

    # BarChart (2-1件数の下)
    chart_anchor_row = LAYOUT['alert_count_row'] + 1 + offset
    _create_alert_chart(wb, ws, f'A{chart_anchor_row}')

    logger.info(f"【総評】填充完成 (alert={alert_count}件, reasons={len(reason_groups)}, offset={offset})")


def _clear_dynamic_area(ws, from_row):
    """清空指定行以下的所有内容和合并单元格"""
    # 取消合并单元格（包括跨越 from_row 的合并区域）
    merged_to_remove = [mr for mr in list(ws.merged_cells.ranges)
                        if mr.max_row >= from_row]
    for mr in merged_to_remove:
        ws.unmerge_cells(str(mr))

    # 强制清理 cell grid 中残留的 MergedCell 对象
    from openpyxl.cell.cell import Cell, MergedCell
    rows_to_clean = range(from_row, ws.max_row + 1)
    cols_to_clean = range(1, ws.max_column + 1)
    for r in rows_to_clean:
        for c in cols_to_clean:
            key = (r, c)
            if key in ws._cells and isinstance(ws._cells[key], MergedCell):
                ws._cells[key] = Cell(ws, row=r, column=c)

    # 清空内容
    max_row = ws.max_row
    max_col = ws.max_column
    for r in range(from_row, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).value = None


def _save_template_section(ws, from_row):
    """保存模板指定行以下的所有单元格值，返回 [(row, col, value), ...]"""
    saved = []
    max_row = ws.max_row
    max_col = ws.max_column
    for r in range(from_row, max_row + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                saved.append((r, c, v))
    return saved


def _restore_template_section(ws, saved_data, start_row, original_start_row,
                              target_year=None, target_month=None):
    """将保存的模板数据写回，行号偏移 = start_row - original_start_row
    同时对 section 7 的月份占位符做动态替换"""
    offset = start_row - original_start_row
    next_month = (target_month % 12) + 1 if target_month else None

    # 先取消目标区域的合并单元格
    if saved_data:
        min_target_row = start_row
        max_target_row = max(orig_r + offset for orig_r, _, _ in saved_data)
        merged_to_remove = [mr for mr in list(ws.merged_cells.ranges)
                            if mr.min_row >= min_target_row and mr.min_row <= max_target_row]
        for mr in merged_to_remove:
            ws.unmerge_cells(str(mr))

    for orig_r, c, v in saved_data:
        new_r = orig_r + offset
        val = v
        if isinstance(val, str) and target_month is not None:
            val = val.replace('M月', f'{target_month}月')
            if next_month is not None:
                val = val.replace('N月', f'{next_month}月')
        try:
            ws.cell(row=new_r, column=c).value = val
        except AttributeError:
            pass  # 跳过 MergedCell


def _is_system_disk(panfu):
    """系统盘：Windows C:\ 或 Linux /dev/vda*"""
    p = str(panfu).strip()
    return p.startswith('/dev/vda') or p.upper().startswith('C:')


def _disk_size_val(row):
    try:
        return float(row.get('磁盘大小(GB)', 0) or 0)
    except (ValueError, TypeError):
        return 0.0


def _ceil_int(val):
    """磁盘大小向上取整（与人工月报一致：99.99->100, 16384.03->16385）"""
    try:
        return int(math.ceil(float(val)))
    except (ValueError, TypeError):
        return val


def _to_c16(val):
    """SVFS C16 最大拡張容量：数字串转 int（与人工月报一致），'-' 等非数字保持原样"""
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return val


def _parse_jst_date(exp_str):
    """资源到期时间(UTC ISO 如 '2026-08-09T16:00:00Z') -> JST 日期 datetime，失败返回 None"""
    if exp_str is None or (isinstance(exp_str, float) and pd.isna(exp_str)) or str(exp_str).strip() in ('', 'nan'):
        return None
    try:
        dt = pd.to_datetime(str(exp_str), utc=True, errors='coerce')
        if pd.isna(dt):
            return None
        dt = dt.tz_convert('Asia/Tokyo').tz_localize(None)
        return datetime(dt.year, dt.month, dt.day)
    except Exception:
        return None


def _find_cpu_row(ws):
    """找【CPU】行(资源表与課金/sections的分界)"""
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value == '【CPU】':
            return r
    return None


def _fill_resource_table_inplace(ws, cmd_data, host_mapping):
    """第一资源表: 定位模板里已有的主机名(B列), 在其行填 CPU/Mem/系统盘/数据盘。
    保留模板合并(B:C主机名, A列環境)。不再清空重建。
    WMS 列映射(与 TSIINFRA 不同, 无 M:N/P:Q 合并):
      D=CPU核 E/F/G=CPU min/max/avg  H=MemGB I/J/K=Mem min/max/avg
      L=系统盘大小 M=系统盘max  N=数据盘大小 O=数据盘max"""
    cpu_row = _find_cpu_row(ws) or (ws.max_row + 1)
    # 按 InstanceName 分组(非SVFS)
    grouped = {}
    order = []
    for _, row in cmd_data.iterrows():
        instance_id = str(row.get('InstanceId', row.get('InstanceName', '')))
        instance_name = str(row.get('InstanceName', ''))
        host_info = host_mapping.get(instance_id, host_mapping.get(instance_name, {}))
        display_name = host_info.get('display_name', instance_name)
        if display_name.upper().startswith('SVFS'):
            continue
        if not host_info and not display_name.lower().startswith('tcp'):
            continue
        if display_name not in grouped:
            grouped[display_name] = []
            order.append(display_name)
        grouped[display_name].append(row)

    filled = 0
    for display_name in order:
        rows = grouped[display_name]
        first = rows[0]
        # 定位主机名(B列, 资源表区 < cpu_row)
        host_row = None
        for r in range(50, cpu_row):
            bv = ws.cell(r, 2).value
            if bv and str(bv).strip() == display_name:
                host_row = r
                break
        if host_row is None:
            logger.warning(f"  【総評】第一表: 未找到主机 {display_name}")
            continue
        system_row = next((rr for rr in rows if _is_system_disk(rr.get('盘符', ''))), rows[0])
        data_candidates = [rr for rr in rows if not _is_system_disk(rr.get('盘符', ''))]
        data_row = max(data_candidates, key=_disk_size_val) if data_candidates else None
        ws.cell(host_row, 4).value = first.get('CPU(核)', '')
        ws.cell(host_row, 5).value = first.get('CPU最小使用率(%)', '')
        ws.cell(host_row, 6).value = first.get('CPU最大使用率(%)', '')
        ws.cell(host_row, 7).value = first.get('CPU平均使用率(%)', '')
        ws.cell(host_row, 8).value = first.get('Memory(GB)', '')
        ws.cell(host_row, 9).value = first.get('内存最小使用率(%)', '')
        ws.cell(host_row, 10).value = first.get('内存最大使用率(%)', '')
        ws.cell(host_row, 11).value = first.get('内存平均使用率(%)', '')
        ws.cell(host_row, 12).value = _ceil_int(system_row.get('磁盘大小(GB)', ''))    # L 系统盘大小
        ws.cell(host_row, 13).value = system_row.get('磁盘最大使用率（%)', '')           # M 系统盘max
        if data_row is not None:
            ws.cell(host_row, 14).value = _ceil_int(data_row.get('磁盘大小(GB)', ''))  # N 数据盘大小
            ws.cell(host_row, 15).value = data_row.get('磁盘最大使用率（%)', '')         # O 数据盘max
        else:
            ws.cell(host_row, 14).value = '-'
            ws.cell(host_row, 15).value = '-'
        filled += 1
    logger.info(f"【総評】第一表(定位填入)完成, {filled}台主机")


def _fill_svfs_table_inplace(ws, cmd_data, svfs_expand_map):
    """第二资源表(SVFS): 定位模板主机名, 主机级 D-M 填首行, 数据盘 N/O/P 每盘一行。
    WMS 无 SVFS 主机 -> 本函数为 no-op (保留与 TSIINFRA 一致的结构)。"""
    cpu_row = _find_cpu_row(ws) or (ws.max_row + 1)
    svfs_hosts = {}
    for _, row in cmd_data.iterrows():
        instance_name = str(row.get('InstanceName', ''))
        if instance_name.upper().startswith('SVFS'):
            disp = instance_name.upper()
            if disp not in svfs_hosts:
                svfs_hosts[disp] = {'iname': instance_name, 'rows': []}
            svfs_hosts[disp]['rows'].append(row)

    filled = 0
    for disp in sorted(svfs_hosts.keys(), reverse=True):   # SVFS13, 12, 11
        info = svfs_hosts[disp]
        iname = info['iname']; rows = info['rows']; first = rows[0]
        host_row = None
        for r in range(50, cpu_row):
            bv = ws.cell(r, 2).value
            if bv and str(bv).strip() == disp:
                host_row = r
                break
        if host_row is None:
            logger.warning(f"  【総評】SVFS表: 未找到 {disp}")
            continue
        # 行跨度(B:C合并: SVFS13=2行, SVFS12=3行, SVFS11=4行)
        span = 1
        for mr in ws.merged_cells.ranges:
            if mr.min_col <= 2 <= mr.max_col and mr.min_row == host_row and mr.max_col >= 3:
                span = mr.max_row - mr.min_row + 1
                break
        system_row = next((rr for rr in rows if _is_system_disk(rr.get('盘符', ''))), rows[0])
        cmd_lookup = {}
        for rr in rows:
            p = str(rr.get('盘符', ''))
            if p and not _is_system_disk(p):
                cmd_lookup[p[0].upper()] = rr
        config_drives = svfs_expand_map.get(iname.lower(), [])
        config_drive_set = {d for d, _ in config_drives}
        for d in cmd_lookup:
            if d not in config_drive_set:
                logger.warning(f"  【SVFS】{disp} 盘符 {d}: cmd有但config无, 忽略")
        drive_list = [(d, _to_c16(v)) for d, v in config_drives] if config_drives \
            else [(d, '-') for d in cmd_lookup.keys()]

        # 主机级 D-M (首行, 跨行合并的top-left: D76:D77..M76:M77 等)
        ws.cell(host_row, 4).value = first.get('CPU(核)', '')
        ws.cell(host_row, 5).value = first.get('CPU最小使用率(%)', '')
        ws.cell(host_row, 6).value = first.get('CPU最大使用率(%)', '')
        ws.cell(host_row, 7).value = first.get('CPU平均使用率(%)', '')
        ws.cell(host_row, 8).value = first.get('Memory(GB)', '')
        ws.cell(host_row, 9).value = first.get('内存最小使用率(%)', '')
        ws.cell(host_row, 10).value = first.get('内存最大使用率(%)', '')
        ws.cell(host_row, 11).value = first.get('内存平均使用率(%)', '')
        ws.cell(host_row, 12).value = _ceil_int(system_row.get('磁盘大小(GB)', ''))   # L
        ws.cell(host_row, 13).value = system_row.get('磁盘最大使用率（%)', '')          # M (M76:M77 merged)
        # 数据盘 N/O/P 每盘一行
        for i, (drive, c16_val) in enumerate(drive_list):
            r = host_row + i
            if i >= span:
                logger.warning(f"  【SVFS】{disp} 数据盘数({len(drive_list)})>模板行跨度({span}), 第{i+1}盘{drive}溢出忽略")
                break
            drow = cmd_lookup.get(drive)
            if drow is not None:
                size = _ceil_int(drow.get('磁盘大小(GB)', ''))
                ws.cell(r, 14).value = f"{size:,}({drive})"            # N
                ws.cell(r, 15).value = drow.get('磁盘剩余大小(GB)', '')  # O
            else:
                ws.cell(r, 14).value = f"({drive})"
            ws.cell(r, 16).value = c16_val                             # P (P:Q merged)
        filled += 1
    logger.info(f"【総評】SVFS表(定位填入)完成, {filled}台主机")


def _fill_rds_table_inplace(ws, rds_op_df, rds_cmd_df, account, tag_filter):
    """RDS 子表(R80-81): 定位模板主机名(B列 DBInstanceDescription), 填 CPU/Mem/Disk。
    模仿 ECS _fill_resource_table_inplace 的定位填入方式, 保留模板合并。
    数据源: aliyun_rds_op.csv(CPU核/memory(GB)/disk(GB)) + aliyun_rds_cmd_summary.csv(使用率)
    RDS 无盘符/系统盘-数据盘区分, 磁盘仅 disk(GB)大小 + 磁盘最大使用率。
    WMS RDS 子表列(合并: L:M=ストレージスペース, N:O=最大使用率):
      D=CPU核 E/F/G=CPU min/max/avg H=memory(GB) I/J/K=Mem min/max/avg
      L=disk(GB)(L:M top-left) N=磁盘最大使用率(N:O top-left)"""
    if rds_op_df is None or rds_op_df.empty:
        logger.info("【総評】RDS表: 无 rds_op 数据, 跳过")
        return
    cpu_row = _find_cpu_row(ws) or (ws.max_row + 1)

    # 过滤 WMS RDS 实例 (账号 + 标签)
    rds_proj = rds_op_df[rds_op_df['资源所属账号'] == account].copy()
    if tag_filter and not rds_proj.empty and '标签' in rds_proj.columns:
        rds_proj = rds_proj[rds_proj['标签'].astype(str).str.contains(re.escape(tag_filter), na=False)]
    if rds_proj.empty:
        logger.info(f"【総評】RDS表: 账号 {account} + 标签 {tag_filter} 无匹配 RDS, 跳过")
        return

    # 合并 cmd_summary 使用率 (by DBInstanceId)
    merged = rds_proj
    if rds_cmd_df is not None and not rds_cmd_df.empty:
        cmd_proj = rds_cmd_df[rds_cmd_df['资源所属账号'] == account].copy()
        merged = rds_proj.merge(
            cmd_proj[['DBInstanceId', 'CPU最小使用率(%)', 'CPU最大使用率(%)', 'CPU平均使用率(%)',
                      '内存最小使用率(%)', '内存最大使用率(%)', '内存平均使用率(%)', '磁盘最大使用率(%)']],
            on='DBInstanceId', how='left'
        )

    filled = 0
    for _, row in merged.iterrows():
        disp = str(row.get('DBInstanceDescription', '')).strip()
        if not disp or disp == 'nan':
            continue
        # 定位 B 列主机名 (资源表区 < cpu_row, RDS 子表在 ECS 表之后)
        host_row = None
        for r in range(50, cpu_row):
            bv = ws.cell(r, 2).value
            if bv and str(bv).strip() == disp:
                host_row = r
                break
        if host_row is None:
            logger.warning(f"  【総評】RDS表: 未找到 {disp}")
            continue
        ws.cell(host_row, 4).value = row.get('CPU(核)', '')             # D CPU核
        ws.cell(host_row, 5).value = row.get('CPU最小使用率(%)', '')     # E
        ws.cell(host_row, 6).value = row.get('CPU最大使用率(%)', '')     # F
        ws.cell(host_row, 7).value = row.get('CPU平均使用率(%)', '')     # G
        ws.cell(host_row, 8).value = row.get('memory(GB)', '')          # H memory(GB) (rds_op 小写列名)
        ws.cell(host_row, 9).value = row.get('内存最小使用率(%)', '')     # I
        ws.cell(host_row, 10).value = row.get('内存最大使用率(%)', '')    # J
        ws.cell(host_row, 11).value = row.get('内存平均使用率(%)', '')    # K
        ws.cell(host_row, 12).value = _ceil_int(row.get('disk(GB)', ''))  # L disk(GB) (L:M top-left)
        ws.cell(host_row, 14).value = row.get('磁盘最大使用率(%)', '')     # N (N:O top-left)
        filled += 1
    logger.info(f"【総評】RDS表(定位填入)完成, {filled}台")


def _replace_section_months(ws, target_year, target_month):
    """section 6 の M月/N月 占位符置換 (6-1 M月の作業状況 -> 当月, 6-2 N月の作業予定 -> 次月)"""
    next_month = (target_month % 12) + 1
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and ('M月' in v or 'N月' in v):
                ws.cell(r, c).value = v.replace('M月', f'{target_month}月').replace('N月', f'{next_month}月')


def _fill_kaiin_renewal_date(ws, renewal_date, c11_map):
    """填充課金表：C7=更新日(作成月第一天), C11=最新有効期限(renewal bill 资源到期时间 JST)
    c11_map: {InstanceName_lower: JST datetime}
    WMS 総評1-8 无 課金一覧 section -> 本函数扫描不到 '課金一覧', 为 no-op。"""
    in_kaiin = False
    filled_c7 = 0
    filled_c11 = 0
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        a_str = str(a).strip() if a is not None else ''
        if '課金一覧' in a_str:
            in_kaiin = True
            continue
        if in_kaiin and a_str.startswith('5、'):
            break
        if in_kaiin:
            c2 = ws.cell(row=r, column=2).value
            c4 = ws.cell(row=r, column=4).value
            if c2 and c4 == '月間サブスクリプション':
                # C7 更新日
                cell7 = ws.cell(row=r, column=7)
                cell7.value = renewal_date
                cell7.number_format = 'yyyy/m/d'
                filled_c7 += 1
                # C11 最新有効期限
                exp = c11_map.get(str(c2).strip().lower())
                if exp:
                    cell11 = ws.cell(row=r, column=11)
                    cell11.value = exp
                    cell11.number_format = 'yyyy/m/d'
                    filled_c11 += 1
    if filled_c7 or filled_c11:
        logger.info(f"【総評】課金填充完成: 更新日{filled_c7}行, 最新有効期限{filled_c11}行")
    else:
        logger.info(f"【総評】課金一覧 未找到 (WMS 模板无課金section), 課金填充跳过")


def _create_alert_chart(wb, summary_ws, anchor_cell):
    """在総評 sheet 创建告警件数 BarChart，引用推移 sheet"""
    if SHEET_TREND not in wb.sheetnames:
        return
    trend_ws = wb[SHEET_TREND]

    # 找到推移sheet中有数据的最后一行(模板 B 列对象名到 sag 止, R22 空即停)
    max_data_row = TREND_DATA_START_ROW
    for r in range(TREND_DATA_START_ROW, trend_ws.max_row + 1):
        if trend_ws.cell(row=r, column=2).value:
            max_data_row = r
        else:
            break

    if max_data_row < TREND_DATA_START_ROW:
        return

    chart = BarChart()
    chart.type = "col"
    chart.title = "全サーバアラート対応件数"
    chart.style = 10
    chart.width = 32
    chart.height = 14

    # C=前月, D=今月
    data_ref = Reference(trend_ws, min_col=3, min_row=TREND_HEADER_ROW,
                         max_col=4, max_row=max_data_row)
    cats_ref = Reference(trend_ws, min_col=2, min_row=TREND_DATA_START_ROW,
                         max_row=max_data_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.shape = 4

    # Y轴固定范围 0-10
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 10
    chart.y_axis.tickLblPos = 'nextTo'
    chart.y_axis.delete = False

    # X轴标签水平（不旋转）
    chart.x_axis.tickLblPos = 'low'
    chart.x_axis.delete = False
    chart.x_axis.txPr = RichText(
        bodyPr=RichTextProperties(rot=0),
        p=[Paragraph(
            pPr=ParagraphProperties(defRPr=CharacterProperties(sz=900)),
            endParaRPr=CharacterProperties(sz=900)
        )]
    )

    # 数据表（图表下方显示数值+系列名称，代替图例）
    chart.plot_area.dTable = DataTable(showHorzBorder=True, showVertBorder=True, showKeys=True)

    # 隐藏独立图例（数据表的showKeys已显示系列名）
    chart.legend = None

    # 绘画区上方留白 10%, 底部留白给数据表(y=0.10, h=0.78 -> 底部约12%)
    from openpyxl.chart.layout import Layout, ManualLayout
    chart.layout = Layout(
        manualLayout=ManualLayout(
            layoutTarget='inner',
            xMode='edge', yMode='edge',
            wMode='edge', hMode='edge',
            x=0.08, y=0.10, w=0.90, h=0.78,
        )
    )

    summary_ws.add_chart(chart, anchor_cell)
    # 强制设置图表尺寸（32cm宽 x 14cm高）
    _set_chart_size(summary_ws._charts[-1], anchor_cell, 32, 14)
    logger.info(f"【総評】BarChart创建完成 at {anchor_cell}")


def populate_trend_sheet(wb, host_display_list, current_alerts, prev_alerts):
    """填充推移 sheet (C=前月, D=今月) -- WMS 定位填入版。
    WMS 模板推移 B 列已有固定对象清单(ECS+RDS+redis+ack+rocketmq+sag), 不能 clear+rewrite
    (会丢失非ECS对象行)。改为: 只对 ECS 主机行填前月/今月告警数, 非ECS对象保持模板原样
    (CEN/SAG/MQ 部分保持模板)。"""
    if SHEET_TREND not in wb.sheetnames:
        return
    ws = wb[SHEET_TREND]
    ecs_names_lower = {str(n).strip().lower() for n in host_display_list}

    filled = 0
    for r in range(TREND_DATA_START_ROW, ws.max_row + 1):
        obj = ws.cell(row=r, column=2).value
        if not obj or not str(obj).strip():
            continue
        obj_str = str(obj).strip()
        if obj_str == '合計':
            break  # R24 合計行(含公式) 及以下第二表 R30-32 不动, 保持模板
        if obj_str.lower() not in ecs_names_lower:
            continue  # 非ECS对象(rds/redis/ack/rocketmq/sag): 保持模板原样

        # 前月告警数 (C列)
        prev_count = 0
        if not prev_alerts.empty:
            prev_count = len(prev_alerts[
                prev_alerts['告警目标'].astype(str).str.lower() == obj_str.lower()
            ])
        ws.cell(row=r, column=3).value = prev_count

        # 今月告警数 (D列)
        current_count = 0
        if not current_alerts.empty:
            current_count = len(current_alerts[
                current_alerts['告警目标'].astype(str).str.lower() == obj_str.lower()
            ])
        ws.cell(row=r, column=4).value = current_count
        filled += 1

    logger.info(f"【推移】定位填入完成，ECS主机{filled}行(非ECS对象保持模板)")


def populate_resource_sheet(wb, sheet_name, metric_type, host_list, data_dir,
                            account, host_mapping, target_year, target_month,
                            display_to_instance):
    """填充リソース sheet (CPU/Memory/Disk) -- WMS 模板タイトル読取版。
    WMS 模板リソース sheet 已有 ■主机名(指標) タイトル(非均一行间隔 16/18/20),
    故不写新タイトル(TSIINFRA 是空白模板写均一18), 改为: 读取模板 ■ タイトル位置,
    在其下放图。监控数据按 display_name -> InstanceName 映射加载(WMS InstanceName=worker-k8s-...
    而 display_name=prod-node1, 两者不同)。
    ScatterChart(数值X轴): 横轴首尾固定(当月1号~月末)5等分(m月d日), 纵轴百分比(min向下取整10倍数,
    max向上取整5倍数), 绘画区上方留白。图尺寸 20.7cm x 7cm(模板 7*20.7cm 注记)。
    每主机独立2列(datetime+average), 避免数据点>行间隔导致重叠覆盖。"""
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]

    from openpyxl.chart import ScatterChart, Series
    from openpyxl.chart.marker import Marker
    from openpyxl.chart.layout import Layout, ManualLayout
    import calendar
    from openpyxl.utils.datetime import to_excel
    metric_label = {'cpu': 'CPU', 'memory': 'Memory', 'disk': 'ディスク'}[metric_type]

    # 横轴范围: 当月(=目标月)1号~月末, 首尾固定后5等分
    last_day = calendar.monthrange(target_year, target_month)[1]
    x_min = to_excel(datetime(target_year, target_month, 1))
    x_max = to_excel(datetime(target_year, target_month, last_day))

    # 读取模板 ■ タイトル行(已预置, 非均一间隔) -> [(title_row, display_name)]
    title_rows = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=2).value
        if isinstance(v, str) and v.startswith('■') and metric_label in v:
            name = v[1:v.index('(')].strip() if '(' in v else v[1:].strip()
            if name:
                title_rows.append((r, name))

    chart_cnt = 0
    for host_idx, (title_row, display_name) in enumerate(title_rows):
        # 每主机独立2列(datetime+average), 避免数据点远超行间隔导致重叠覆盖
        data_col = RESOURCE_DATA_START_COL + host_idx * 2
        instance_name = display_to_instance.get(display_name, display_name)

        # 加载监控数据(按 InstanceName, WMS worker-k8s-... 文件名)
        monitor_df = load_monitoring_data(data_dir, account, instance_name, metric_type)

        if not monitor_df.empty and 'datetime' in monitor_df.columns:
            header_row = title_row
            ws.cell(row=header_row, column=data_col).value = 'datetime'
            ws.cell(row=header_row, column=data_col + 1).value = 'average'

            if metric_type == 'disk' and '盘符' in monitor_df.columns:
                disk_groups = monitor_df.groupby('盘符')['average'].mean()
                if not disk_groups.empty:
                    monitor_df = monitor_df[monitor_df['盘符'] == disk_groups.idxmax()]

            num_points = len(monitor_df)
            # 写真实 datetime(数值X轴读为日期序列号) + average 数值；不清列(清列会让图表数据源失效)
            for pt_idx, (_, pt) in enumerate(monitor_df.iterrows()):
                r = header_row + 1 + pt_idx
                dt_cell = ws.cell(row=r, column=data_col)
                dt_cell.value = pt['datetime']
                dt_cell.number_format = 'm月d日'
                ws.cell(row=r, column=data_col + 1).value = pt.get('average', 0)

            chart = ScatterChart()
            chart.title = f"{display_name} ({metric_label}使用率)"
            chart.style = 10
            chart.width = RESOURCE_CHART_W
            chart.height = RESOURCE_CHART_H
            # 纵轴: 百分比刻度; min=数据最小值向下取整到10的倍数, max=数据最大值向上取整到5的倍数
            avg_vals = pd.to_numeric(monitor_df['average'], errors='coerce')
            avg_max = 0.0 if pd.isna(avg_vals.max()) else float(avg_vals.max())
            avg_min = 0.0 if pd.isna(avg_vals.min()) else float(avg_vals.min())
            y_max = max(5, math.ceil(avg_max / 5) * 5)
            y_min = math.floor(avg_min / 10) * 10
            if y_max <= y_min:
                y_max = y_min + 5
            chart.y_axis.number_format = '0"%"'
            chart.y_axis.scaling.min = y_min
            chart.y_axis.scaling.max = y_max
            chart.y_axis.majorUnit = 5
            chart.y_axis.tickLblPos = 'nextTo'
            chart.y_axis.delete = False
            # 横轴: 数值轴(日期序列号), 首尾固定+固定5等分, 无轴标题
            N_SEGMENTS = 5
            x_unit = (x_max - x_min) / N_SEGMENTS * (1 + 1e-9)
            chart.x_axis.number_format = 'm月d日'
            chart.x_axis.scaling.min = x_min
            chart.x_axis.scaling.max = x_min + N_SEGMENTS * x_unit
            chart.x_axis.majorUnit = x_unit
            chart.x_axis.majorGridlines = None  # 关闭竖向网格线
            chart.x_axis.tickLblPos = 'low'
            chart.x_axis.delete = False
            chart.legend = None
            chart.layout = Layout(
                manualLayout=ManualLayout(
                    layoutTarget='inner',
                    xMode='edge', yMode='edge',
                    wMode='edge', hMode='edge',
                    x=0.1, y=0.10, w=0.85, h=0.85,
                )
            )

            # xvalues(日期, 不含标题) + yvalues(average, 含标题)
            xvalues = Reference(ws, min_col=data_col, min_row=header_row + 1,
                                max_row=header_row + num_points)
            yvalues = Reference(ws, min_col=data_col + 1, min_row=header_row,
                                max_row=header_row + num_points)
            series = Series(yvalues, xvalues, title_from_data=True)
            series.marker = Marker(symbol='none')
            series.graphicalProperties = GraphicalProperties()
            series.graphicalProperties.line = LineProperties(
                solidFill="4472C4",
                w=12700
            )
            chart.series.append(series)

            ws.add_chart(chart, f"B{title_row + 1}")
            _set_chart_size(ws._charts[-1], f"B{title_row + 1}", RESOURCE_CHART_W, RESOURCE_CHART_H)
            chart_cnt += 1
        else:
            logger.warning(f"  [{sheet_name}] {display_name}(InstanceName={instance_name}): 无监控数据")

    logger.info(f"【{sheet_name}】填充完成，模板タイトル{len(title_rows)}个, 图表{chart_cnt}个")


def populate_rds_sheet(wb, sheet_name, data_dir, account, target_year, target_month):
    """填充 RDS sheet (9-4) CPU/Memory/Disk 折线图 -- 模仿 ECS populate_resource_sheet。
    RDS sheet 把三个指标放一个 sheet, ■タイトル含指标名((CPU使用率)/(Memory使用率)/(ディスク使用率)),
    间隔均一18行。读取模板 ■ タイトル位置, 解析 host + metric, 加载 rds_{metric}_{host}.csv 放图。
    监控文件名带 rds_ 前缀(区别于 ECS 的 cpu_/memory_/disk_)。
    ScatterChart(数值X轴) + 图尺寸 20.7x7cm(模板注記), 与 ECS リソース图同一方式。"""
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]

    from openpyxl.chart import ScatterChart, Series
    from openpyxl.chart.marker import Marker
    from openpyxl.chart.layout import Layout, ManualLayout
    import calendar
    from openpyxl.utils.datetime import to_excel

    # 标题内指标 -> metric_key; metric_key -> 图表用 label
    metric_map = [('CPU', 'cpu'), ('Memory', 'memory'), ('ディスク', 'disk')]
    metric_label_map = {'cpu': 'CPU', 'memory': 'Memory', 'disk': 'ディスク'}

    # 横轴范围: 当月1号~月末, 首尾固定后5等分
    last_day = calendar.monthrange(target_year, target_month)[1]
    x_min = to_excel(datetime(target_year, target_month, 1))
    x_max = to_excel(datetime(target_year, target_month, last_day))

    # 读取模板 ■ タイトル行 -> [(title_row, host_name, metric_key)]
    title_rows = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=2).value
        if not isinstance(v, str) or not v.startswith('■') or '使用率' not in v:
            continue
        name = v[1:v.index('(')].strip() if '(' in v else v[1:].strip()
        metric_key = next((key for label, key in metric_map if label in v), None)
        if name and metric_key:
            title_rows.append((r, name, metric_key))

    chart_cnt = 0
    for host_idx, (title_row, display_name, metric_key) in enumerate(title_rows):
        # 每标题独立2列(datetime+average), 避免数据点>行间隔导致重叠覆盖
        data_col = RESOURCE_DATA_START_COL + host_idx * 2
        metric_label = metric_label_map[metric_key]

        monitor_df = load_rds_monitoring_data(data_dir, account, display_name, metric_key)

        if not monitor_df.empty and 'datetime' in monitor_df.columns:
            header_row = title_row
            ws.cell(row=header_row, column=data_col).value = 'datetime'
            ws.cell(row=header_row, column=data_col + 1).value = 'average'

            num_points = len(monitor_df)
            for pt_idx, (_, pt) in enumerate(monitor_df.iterrows()):
                r = header_row + 1 + pt_idx
                dt_cell = ws.cell(row=r, column=data_col)
                dt_cell.value = pt['datetime']
                dt_cell.number_format = 'm月d日'
                ws.cell(row=r, column=data_col + 1).value = pt.get('average', 0)

            chart = ScatterChart()
            chart.title = f"{display_name} ({metric_label}使用率)"
            chart.style = 10
            chart.width = RESOURCE_CHART_W
            chart.height = RESOURCE_CHART_H
            # 纵轴: 百分比; min向下取整10倍数, max向上取整5倍数
            avg_vals = pd.to_numeric(monitor_df['average'], errors='coerce')
            avg_max = 0.0 if pd.isna(avg_vals.max()) else float(avg_vals.max())
            avg_min = 0.0 if pd.isna(avg_vals.min()) else float(avg_vals.min())
            y_max = max(5, math.ceil(avg_max / 5) * 5)
            y_min = math.floor(avg_min / 10) * 10
            if y_max <= y_min:
                y_max = y_min + 5
            chart.y_axis.number_format = '0"%"'
            chart.y_axis.scaling.min = y_min
            chart.y_axis.scaling.max = y_max
            chart.y_axis.majorUnit = 5
            chart.y_axis.tickLblPos = 'nextTo'
            chart.y_axis.delete = False
            # 横轴: 数值轴(日期序列号), 首尾固定+5等分, 无标题, 关竖向网格
            N_SEGMENTS = 5
            x_unit = (x_max - x_min) / N_SEGMENTS * (1 + 1e-9)
            chart.x_axis.number_format = 'm月d日'
            chart.x_axis.scaling.min = x_min
            chart.x_axis.scaling.max = x_min + N_SEGMENTS * x_unit
            chart.x_axis.majorUnit = x_unit
            chart.x_axis.majorGridlines = None
            chart.x_axis.tickLblPos = 'low'
            chart.x_axis.delete = False
            chart.legend = None
            chart.layout = Layout(
                manualLayout=ManualLayout(
                    layoutTarget='inner',
                    xMode='edge', yMode='edge',
                    wMode='edge', hMode='edge',
                    x=0.1, y=0.10, w=0.85, h=0.85,
                )
            )

            xvalues = Reference(ws, min_col=data_col, min_row=header_row + 1,
                                max_row=header_row + num_points)
            yvalues = Reference(ws, min_col=data_col + 1, min_row=header_row,
                                max_row=header_row + num_points)
            series = Series(yvalues, xvalues, title_from_data=True)
            series.marker = Marker(symbol='none')
            series.graphicalProperties = GraphicalProperties()
            series.graphicalProperties.line = LineProperties(
                solidFill="4472C4",
                w=12700
            )
            chart.series.append(series)

            ws.add_chart(chart, f"B{title_row + 1}")
            _set_chart_size(ws._charts[-1], f"B{title_row + 1}", RESOURCE_CHART_W, RESOURCE_CHART_H)
            chart_cnt += 1
        else:
            logger.warning(f"  [{sheet_name}] {display_name}({metric_label}): 无监控数据")

    logger.info(f"【{sheet_name}】填充完成，模板タイトル{len(title_rows)}个, 图表{chart_cnt}个")


def add_above_marker(wb):
    """[TSIINFRA 用] 在最后一个 B 列含 ■ 的 sheet 末尾 M 列添加 "以上"。
    WMS 不调用此函数: WMS 模板 MQ sheet 末尾(R40)已有 "以上"(=全リソース終端标记),
    且 reversed-find 会误中 MQ sheet(在 Disk 之后)。WMS 保持模板 MQ 的 "以上" 不动。
    保留函数定义以与 TSIINFRA 对齐。"""
    target_ws = None
    anchor_row = None
    for ws in reversed(wb.worksheets):
        last_bullet = None
        for cell in ws['B']:
            if isinstance(cell.value, str) and '■' in cell.value:
                last_bullet = cell.row
        if last_bullet is not None:
            target_ws, anchor_row = ws, last_bullet
            break
    if target_ws is None or anchor_row is None:
        logger.warning("「以上」标记: 未找到 B 列含 ■ 的锚点 sheet, 跳过")
        return
    target_row = anchor_row + RESOURCE_INTERVAL_ROWS
    above_cell = target_ws.cell(row=target_row, column=13)  # M 列 = 13
    above_cell.value = "以上"
    above_cell.font = Font(name='MS PGothic', size=12)
    logger.info(f"「以上」标记: sheet={target_ws.title} 锚点 B{anchor_row} -> M{target_row}")


# --- V. 主逻辑 ---
def build_project_report(data, target_year, target_month, today):
    """为 TSIWMS 项目生成运维月报"""
    project_key = PROJECT_KEY

    # 名称映射
    names_df = data.get('names', pd.DataFrame())
    name_row = names_df[names_df['Project Key'] == project_key] if not names_df.empty else pd.DataFrame()
    if not name_row.empty:
        company_name = str(name_row.iloc[0]['公司名'])
        system_name = str(name_row.iloc[0]['系统名'])
    else:
        logger.warning(f"项目 {project_key} 无名称映射，跳过")
        return None

    logger.info(f"\n{'='*60}")
    logger.info(f"开始处理项目: {project_key} ({company_name})")
    logger.info(f"{'='*60}")

    # 主机映射
    host_mapping = get_host_mapping(data.get('hosts', pd.DataFrame()), project_key)

    # 账号和标签
    account_df = data['op_account']
    proj_accounts = account_df[account_df['Project Key'] == project_key]
    if proj_accounts.empty:
        logger.warning(f"项目 {project_key} 无对应账号，跳过")
        return None

    account_row = proj_accounts.iloc[0]
    account = account_row['资源所属账号']
    tag_filter = parse_account_tag(account_row.get('标签', ''))
    logger.info(f"账号: {account}, 标签过滤: {tag_filter}")

    # ECS实例
    instances_df = get_project_instances(account, tag_filter, data['ecs'])
    if instances_df.empty:
        logger.warning(f"项目 {project_key} 无匹配ECS实例，跳过")
        return None

    # 去重
    instances_df = instances_df.drop_duplicates(subset=['InstanceId'], keep='first')
    host_list = instances_df['InstanceName'].tolist()
    logger.info(f"匹配到 {len(host_list)} 台ECS实例(去重后)")

    # display_name -> InstanceName 映射 (WMS: display_name=prod-node1, InstanceName=worker-k8s-...)
    display_to_instance = {}
    for _, inst_row in instances_df.iterrows():
        iid = str(inst_row['InstanceId'])
        iname = str(inst_row['InstanceName'])
        info = host_mapping.get(iid, {})
        disp = info.get('display_name', iname)
        display_to_instance[disp] = iname

    # 资源使用率汇总
    cmd_df = data['cmd_summary']
    cmd_proj = cmd_df[cmd_df['资源所属账号'] == account].copy()
    cmd_proj = cmd_proj[cmd_proj['InstanceName'].isin(host_list)]

    if not cmd_proj.empty:
        ecs_info = instances_df[['InstanceName', 'InstanceId', 'CPU(核)', 'Memory(MB)']].copy()
        ecs_info['Memory(GB)'] = (ecs_info['Memory(MB)'] / 1024).round(0).astype(int)
        # cmd_summary 已有 InstanceId, 只补 CPU(核)/Memory(GB)。先按 InstanceName 去重 ecs_info,
        # 避免 WMS k8s worker 节点(4prod/4uat 共享同一 InstanceName) merge 时行膨胀 +
        # InstanceId_x/_y 列冲突(会使 row.get('InstanceId') 失效 -> host_mapping 查不到 -> 显示名回退到长 InstanceName)
        ecs_info_dedup = ecs_info.drop_duplicates(subset=['InstanceName'], keep='first')
        cmd_proj = cmd_proj.merge(
            ecs_info_dedup[['InstanceName', 'CPU(核)', 'Memory(GB)']],
            on='InstanceName', how='left'
        )
        # 按 (InstanceId, 盘符) 去重(非 InstanceName): k8s 节点共享 InstanceName 但 InstanceId 唯一,
        # 按 InstanceName 去重会把 4 个 prod-node 折叠成 1。保留每节点每盘符一行
        cmd_proj = cmd_proj.drop_duplicates(subset=['InstanceId', '盘符'], keep='first')

        # 按环境映射排序
        if host_mapping:
            env_order = {}
            for iid, info in host_mapping.items():
                env_order[info['display_name']] = list(host_mapping.keys()).index(iid)
            cmd_proj['sort_key'] = cmd_proj['InstanceName'].map(
                lambda x: env_order.get(x, 999)
            )
            cmd_proj = cmd_proj.sort_values('sort_key').drop(columns=['sort_key'])

    # 告警数据
    current_alerts, prev_alerts = get_alerts_for_project(
        data['alert'], project_key, target_year, target_month
    )
    logger.info(f"告警(过滤后): 今月{len(current_alerts)}件, 前月{len(prev_alerts)}件")

    # 加载模板
    template_path = data['template_path']
    logger.info(f"使用模板: {os.path.basename(template_path)}")
    wb = load_workbook(template_path)

    # 填充各 Sheet
    populate_cover(wb, company_name, system_name, target_year, target_month, today)

    # 推移 sheet（使用显示名，过滤+排序与人工版本一致）
    # 排序规则：本番tcp -> 開発tcp -> 本番SV(非SVFS) -> SVFS
    mapping_keys = list(host_mapping.keys())  # 保持配置顺序
    trend_items = []
    for _, inst_row in instances_df.iterrows():
        iid = str(inst_row['InstanceId'])
        iname = str(inst_row['InstanceName'])
        info = host_mapping.get(iid, host_mapping.get(iname, {}))
        display_name = info.get('display_name', iname)
        env = info.get('env', '')

        # 过滤：排除不在 host_mapping 且不以 tcp 开头的主机
        if not info and not display_name.lower().startswith('tcp'):
            if not display_name.upper().startswith('SVFS'):
                continue
        # 不在 mapping 的 tcp 主机默认開発環境
        if not env and display_name.lower().startswith('tcp'):
            env = '開発環境'

        # 排序 key: (group, order_within_group)
        map_idx = mapping_keys.index(iid) if iid in mapping_keys else (
            mapping_keys.index(iname) if iname in mapping_keys else 999)
        if display_name.upper().startswith('SVFS'):
            display_name = display_name.upper()  # SVFS 统一大写
            sort_key = (3, [-ord(c) for c in display_name])  # SVFS 降序
        elif display_name.lower().startswith('tcp') and env == '本番環境':
            sort_key = (0, map_idx)       # 本番 tcp
        elif display_name.lower().startswith('tcp'):
            sort_key = (1, map_idx)       # 開発 tcp
        else:
            sort_key = (2, map_idx)       # 本番 SV (SVFR, SVVS)
        trend_items.append((sort_key, display_name))

    trend_items.sort(key=lambda x: x[0])
    trend_display_list = [name for _, name in trend_items]

    populate_trend_sheet(wb, trend_display_list, current_alerts, prev_alerts)

    # 課金 C11 最新有効期限: renewal bill 资源到期时间(JST) 按 InstanceId 匹配
    # (WMS 无 課金一覧, c11_map 实际不使用, 保留加载逻辑与 TSIINFRA 一致)
    bill_path = os.path.join(data['data_dir'],
                             f'aliyun_renewal_bill_{target_year}-{target_month:02d}.csv')
    c11_map = {}  # InstanceName_lower -> JST datetime
    if os.path.exists(bill_path):
        try:
            bill_df = safe_read_csv(bill_path)
            bill_exp = {}
            for _, br in bill_df.iterrows():
                iid = str(br.get('资源id', '')).strip()
                if iid and iid != 'nan':
                    bill_exp[iid] = br.get('资源到期时间', '')
            for _, er in data['ecs'].iterrows():
                iname = str(er.get('InstanceName', '')).strip().lower()
                iid = str(er.get('InstanceId', '')).strip()
                if iname and iid and iid in bill_exp:
                    d = _parse_jst_date(bill_exp[iid])
                    if d:
                        c11_map[iname] = d
            logger.info(f"課金C11最新有効期限: 加载{len(c11_map)}台 ({os.path.basename(bill_path)})")
        except Exception as e:
            logger.warning(f"加载renewal bill失败: {e}")
    else:
        logger.warning(f"renewal bill不存在: {bill_path} (課金C11将留空)")

    populate_summary(wb, current_alerts, prev_alerts, cmd_proj,
                     target_year, target_month, host_mapping, today,
                     data.get('svfs_expand', {}), c11_map,
                     data.get('rds_op', pd.DataFrame()), data.get('rds_cmd', pd.DataFrame()),
                     account, tag_filter)

    # リソース sheets (读取模板 ■ タイトル位置放图, 按显示名->InstanceName 加载监控数据)
    populate_resource_sheet(wb, SHEET_CPU, 'cpu', host_list, data['data_dir'],
                            account, host_mapping, target_year, target_month,
                            display_to_instance)
    populate_resource_sheet(wb, SHEET_MEMORY, 'memory', host_list, data['data_dir'],
                            account, host_mapping, target_year, target_month,
                            display_to_instance)
    populate_resource_sheet(wb, SHEET_DISK, 'disk', host_list, data['data_dir'],
                            account, host_mapping, target_year, target_month,
                            display_to_instance)

    # RDS sheet (9-4): CPU/Memory/Disk 折线图, 模仿 ECS リソース图, 读模板 ■ タイトル放图
    populate_rds_sheet(wb, SHEET_RDS, data['data_dir'], account, target_year, target_month)

    # WMS 不调用 add_above_marker: 模板 MQ sheet 末尾已有 "以上"(全リソース終端标记),
    # CEN/SAG/MQ 部分保持模板原样。ECS Disk sheet 不另加 "以上"。

    # 保存
    output_dir = data['output_dir']
    project_folder = os.path.join(output_dir, project_key)
    os.makedirs(project_folder, exist_ok=True)

    year_month_str = f"{target_year}{target_month:02d}"
    output_filename = f"【{company_name}様向け】WMS月次報告書({year_month_str}).xlsx"
    output_path = os.path.join(project_folder, output_filename)

    wb.save(output_path)
    wb.close()

    # 修正图表尺寸（openpyxl 保存时不正确写入 anchor ext）
    # drawing1=総評 BarChart(32x14), drawing2-4=リソース LineCharts, drawing5=RDS(9-4), 均 20.7x7
    _fix_chart_sizes_in_xlsx(output_path, {
        1: (32, 14),                              # 総評 BarChart: 32cm x 14cm
        2: (RESOURCE_CHART_W, RESOURCE_CHART_H),  # リソース (CPU)
        3: (RESOURCE_CHART_W, RESOURCE_CHART_H),  # リソース (Memory)
        4: (RESOURCE_CHART_W, RESOURCE_CHART_H),  # リソース (Disk)
        5: (RESOURCE_CHART_W, RESOURCE_CHART_H),  # RDS (9-4)
    })

    logger.info(f"报表生成成功: {output_path}")
    return output_path


def main():
    logger.info("=" * 80)
    logger.info(f"开始执行 TSIWMS 运维月报生成: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    try:
        today = datetime.today()
        target_date = today.replace(day=1) - timedelta(days=1)
        target_year = target_date.year
        target_month = target_date.month
        logger.info(f"目标月份: {target_year}年{target_month}月")

        script_dir = os.path.dirname(os.path.abspath(__file__))
        data = load_all_data(script_dir)

        result = build_project_report(data, target_year, target_month, today)
        if result:
            logger.info(f"\n{'='*80}")
            logger.info(f"TSIWMS 报表生成成功: {result}")
        else:
            logger.error("TSIWMS 报表生成失败")

        logger.info(f"结束时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    except Exception as e:
        logger.error(f"程序执行失败: {str(e)}", exc_info=True)


if __name__ == "__main__":
    main()
