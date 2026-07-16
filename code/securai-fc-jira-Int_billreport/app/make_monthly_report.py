import os
import re
import glob
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from datetime import datetime
from dateutil.relativedelta import relativedelta
import logging
import math

# --- I. 常量定义 ---
SHEET_COVER = '表紙'
SHEET_SUMMARY = '総評'

# 【表紙】Sheet 常量
COVER_CREATE_DATE_ROW = 34  # D34 作成日单元格
COVER_CREATE_DATE_COL = 4  # D列

# 【総評】Sheet 基础常量
SUMMARY_RATE_ROW = 53  # 汇率基准行
SUMMARY_RATE_COL = 4  # D列
SUMMARY_ACCOUNT_ROW = 17
SUMMARY_ACCOUNT_COL = 8  # H列
SUMMARY_DETAIL_START_ROW = 59  # 消费明细插入行
SUMMARY_DETAIL_COLS = {
    'ProductCode': 7,  # G列
    '资源id': 22,  # V列
    '消费金额': 37,  # AK列
}
SUMMARY_MERGE_RANGES = [
    (7, 21),  # G-U
    (22, 36),  # V-AJ
    (37, 49),  # AK-AW
]

# 【総評】Sheet 12个月明细填充区域（核心修改：仅偶数行）
# D28-AK50 范围内的偶数行：28/30/32/34/36/38/40/42/44/46/48/50，正好12行对应12个月
SUMMARY_MONTHLY_TARGET_ROWS = list(range(28, 51, 2))
SUMMARY_MONTHLY_COL_OFFSET = 3  # 列偏移：原A列→D列，+3
# 列映射保持不变，完全对齐D28-AK50区域
MONTHLY_COL_YEAR_MONTH = 1 + SUMMARY_MONTHLY_COL_OFFSET  # 年月 → D列
MONTHLY_COL_AMOUNT_USD = 4 + SUMMARY_MONTHLY_COL_OFFSET  # 消费金额USD → G列
MONTHLY_COL_AMOUNT_JPY = 11 + SUMMARY_MONTHLY_COL_OFFSET  # 消费金额JPY → N列
MONTHLY_COL_TAX = 18 + SUMMARY_MONTHLY_COL_OFFSET  # 税金 → U列
MONTHLY_COL_AGENT_FEE = 26 + SUMMARY_MONTHLY_COL_OFFSET  # 代行费用 → AC列
MONTHLY_COL_FINAL = 34 + SUMMARY_MONTHLY_COL_OFFSET  # 最终金额 → AK列


# --- II. 辅助函数 ---
def init_logger():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    log_dir = os.path.join(script_dir, '/jira', 'log')
    os.makedirs(log_dir, exist_ok=True)
    log_file = os.path.join(log_dir, 'monthly_report.log')

    logging.basicConfig(
        level=logging.DEBUG,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler()
        ]
    )
    return logging.getLogger(__name__)


logger = init_logger()


def safe_read_csv(file_path):
    for encoding in ['utf-8', 'gbk', 'shift-jis', 'latin1']:
        try:
            df = pd.read_csv(file_path, encoding=encoding, sep=',')
            logger.debug(f"成功使用 {encoding} 编码读取文件: {file_path}")
            return df
        except UnicodeDecodeError:
            continue
        except Exception as e:
            logger.warning(f"使用 {encoding} 读取文件 {file_path} 失败: {e}")
            continue

    error_msg = f"无法使用 utf-8/gbk/shift-jis/latin1 编码读取文件: {file_path}"
    logger.error(error_msg)
    raise IOError(error_msg)


def safe_write_cell(sheet, row_num, column, value):
    try:
        sheet.cell(row=row_num, column=column).value = value
    except AttributeError:
        coord = sheet.cell(row=row_num, column=column).coordinate
        logger.warning(f"跳过写入单元格 {coord} (R{row_num}C{column})")
        pass


# --- III. 主要逻辑函数 ---
def load_and_preprocess_data(script_dir, target_date):
    def get_absolute_path(rel_path):
        return os.path.normpath(os.path.join(script_dir, rel_path))

    template_path = get_absolute_path("../../jira/monthly_report/课金月报模板.xlsx")
    resale_csv = get_absolute_path("../../jira/data/resale_business_details.csv")
    jira_csv = get_absolute_path("../../jira/data/jira_get_account.csv")
    aliyun_bill_dir = get_absolute_path("../../jira/data")
    aliyun_bill_pattern = os.path.join(aliyun_bill_dir, "aliyun_bill_*.csv")
    rate_csv = get_absolute_path("../../jira/data/rate.csv")
    output_dir = get_absolute_path("../../jira/monthly_report")
    os.makedirs(output_dir, exist_ok=True)

    logger.info(f"模板文件路径：{template_path}")
    for path in [template_path, resale_csv, jira_csv, rate_csv]:
        if not os.path.exists(path):
            error_msg = f"关键文件不存在：{path}"
            logger.error(error_msg)
            raise FileNotFoundError(error_msg)

    # 读取汇率
    logger.info("读取汇率数据...")
    rate_df = safe_read_csv(rate_csv)
    rate_df['年月'] = rate_df['年月'].astype(str).str.strip()
    rate_dict = pd.Series(rate_df['汇率'].values, index=rate_df['年月']).to_dict()
    logger.info(f"汇率加载完成: {rate_dict}")

    # 读取转售数据
    logger.info("开始读取转售数据...")
    resale_df = safe_read_csv(resale_csv)
    resale_df['原始代行费率'] = resale_df['代行费率'].astype(str).replace('nan', '', regex=False).fillna('')
    if '税率' in resale_df.columns:
        resale_df['原始税率'] = resale_df['税率'].astype(str).replace('nan', '', regex=False).fillna('')
    else:
        resale_df['原始税率'] = ''

    resale_df['代行费率'] = resale_df['代行费率'].apply(
        lambda x: float(str(x).strip().strip('%')) / 100 if pd.notna(x) and isinstance(x, str) and str(
            x).strip().endswith('%') else pd.to_numeric(x, errors='coerce') if pd.notna(x) else 0.0
    ).fillna(0.0)

    if '税率' not in resale_df.columns:
        logger.warning("转售数据中未找到 '税率' 列，使用默认值 0.1")
        resale_df['税率'] = 0.1
    resale_df['税率'] = resale_df['税率'].apply(
        lambda x: float(str(x).strip().strip('%')) / 100 if pd.notna(x) and isinstance(x, str) and str(
            x).strip().endswith('%') else pd.to_numeric(x, errors='coerce') if pd.notna(x) else 0.1
    ).fillna(0.1)

    resale_df['客户公司名称'] = resale_df['Project Name'].apply(
        lambda x: (re.search(r'【(.*?)】', str(x)).group(1) if re.search(r'【(.*?)】', str(x)) else str(x)) + '   御中'
    )

    # 合并Jira数据
    logger.info("读取Jira账号数据并合并...")
    jira_df = safe_read_csv(jira_csv)
    jira_df['标签'] = jira_df['标签'].astype(str).replace('nan', '', regex=False).str.strip()
    merged_df = pd.merge(resale_df, jira_df, on='Project Key', how='left')
    merged_df['资源所属账号'] = merged_df['资源所属账号'].fillna('未知账号')
    merged_df['Project Key'] = merged_df['Project Key'].fillna('未知项目')
    merged_df = merged_df[merged_df['资源所属账号'] != '未知账号']
    logger.info(f"数据合并完成，共 {len(merged_df)} 条记录")

    # 读取阿里云账单
    logger.info("读取阿里云账单数据...")
    last_month_year = target_date.year
    last_month = target_date.month
    end_period = f"{last_month_year}-{last_month:02d}"
    start_date = datetime(last_month_year, last_month, 1) - relativedelta(months=11)
    start_period = f"{start_date.year}-{start_date.month:02d}"

    aliyun_bills_by_period = {}
    bill_files = glob.glob(aliyun_bill_pattern)

    for bill_file in bill_files:
        basename = os.path.basename(bill_file)
        match = re.search(r'aliyun_bill_(\d{4})-?(\d{2})\.csv', basename)
        if match:
            period = f"{match.group(1)}-{match.group(2)}"
            if start_period <= period <= end_period:
                try:
                    df = safe_read_csv(bill_file)
                except IOError:
                    logger.warning(f"无法读取账单文件 {basename}，跳过。")
                    continue

                required_cols = ['资源所属账号', '标签', '消费金额', 'ProductCode', '资源id']
                if not all(col in df.columns for col in required_cols):
                    logger.warning(f"账单 {period} 缺少必要列，跳过")
                    continue

                df['资源所属账号'] = df['资源所属账号'].astype(str).str.strip()
                df['标签'] = df['标签'].astype(str).replace('nan', '', regex=False).str.strip()
                df['消费金额'] = pd.to_numeric(df['消费金额'], errors='coerce').fillna(0)
                aliyun_bills_by_period[period] = df
                logger.info(f"账单 {period} 加载完成，共 {len(df)} 条记录")
        else:
            logger.warning(f"跳过不符合命名规范的文件：{basename}")

    if not aliyun_bills_by_period:
        logger.warning("未找到符合周期的阿里云账单数据")

    return template_path, output_dir, merged_df, aliyun_bills_by_period, start_period, end_period, rate_dict


def process_account_bills(account, current_tag, aliyun_bills_by_period, start_period, end_period):
    current_tag = str(current_tag).strip() if pd.notna(current_tag) else ""
    monthly_summary = {}
    monthly_details = {}

    current_date = datetime.strptime(start_period, "%Y-%m")
    end_date = datetime.strptime(end_period, "%Y-%m")
    period_list = []

    while current_date <= end_date:
        period_list.append(current_date.strftime("%Y-%m"))
        current_date += relativedelta(months=1)

    for period in period_list:
        full_df = aliyun_bills_by_period.get(period)
        if full_df is None:
            monthly_summary[period] = 0.0
            monthly_details[period] = pd.DataFrame()
            continue

        account_df = full_df.query('`资源所属账号` == @account')
        if current_tag:
            filtered_df = account_df.query('`标签` == @current_tag')
        else:
            filtered_df = account_df

        if not filtered_df.empty:
            total_amount_period = filtered_df['消费金额'].sum()
            monthly_summary[period] = total_amount_period
            monthly_details[period] = filtered_df[['资源所属账号', 'ProductCode', '资源id', '消费金额']]
        else:
            monthly_summary[period] = 0.0
            monthly_details[period] = pd.DataFrame()

        logger.debug(f"周期 {period} 筛选后金额: {monthly_summary[period]:.2f}")

    last_month_key = end_period
    total_amount = monthly_summary.get(last_month_key, 0.0)
    account_details = monthly_details.get(last_month_key, pd.DataFrame())

    return monthly_summary, total_amount, account_details


def populate_template(template_path, row_data, monthly_summary, account_details, date_range, date_str_d58, rate_dict,
                      end_period, today):
    account = row_data['资源所属账号']
    agent_rate = row_data.get('代行费率', 0.0)
    tax_rate = row_data.get('税率', 0.1)
    current_tag = row_data.get('标签')
    company_name = row_data.get('客户公司名称', '未知公司')

    original_tax_rate_str = row_data.get('原始税率', '')
    original_agent_rate_str = row_data.get('原始代行费率', '')
    replace_tax_val = original_tax_rate_str.strip().replace('nan', '')
    replace_agent_val = original_agent_rate_str.strip().replace('nan', '')

    try:
        current_wb = load_workbook(template_path)
    except Exception as e:
        logger.error(f"加载模板失败：{str(e)}", exc_info=True)
        raise

    # 填充【表紙】Sheet
    if SHEET_COVER in current_wb.sheetnames:
        cover_sheet = current_wb[SHEET_COVER]
        for cell in cover_sheet.iter_rows(values_only=False):
            for c in cell:
                if isinstance(c.value, str):
                    if '年月日' in c.value:
                        c.value = c.value.replace('年月日', date_range)
                    if '客户公司名称' in c.value:
                        c.value = c.value.replace('客户公司名称', company_name)

        # D34 作成日填充
        create_date_str = f"作成日：{today.year}/{today.month}/{today.day}"
        safe_write_cell(cover_sheet, COVER_CREATE_DATE_ROW, COVER_CREATE_DATE_COL, create_date_str)
        logger.debug(f"【{SHEET_COVER}】D34 作成日填充完成：{create_date_str}")

    # 填充【総評】Sheet
    if SHEET_SUMMARY in current_wb.sheetnames:
        summary_sheet = current_wb[SHEET_SUMMARY]

        # 全局占位符 AAA/BBB 替换
        logger.info(f"【{SHEET_SUMMARY}】全局替换 AAA/BBB 占位符")
        for row in summary_sheet.iter_rows(values_only=False):
            for cell in row:
                if isinstance(cell.value, str):
                    if 'AAA' in cell.value:
                        cell.value = cell.value.replace('AAA', replace_tax_val)
                    elif 'BBB' in cell.value:
                        cell.value = cell.value.replace('BBB', replace_agent_val)

        # D53 汇率基准填充
        display_rate_value = row_data.get('汇率', 0.0) * 100
        rate_content = f"レート基準 {date_str_d58} 100 USD= {display_rate_value:.1f} JPY"
        safe_write_cell(summary_sheet, SUMMARY_RATE_ROW, SUMMARY_RATE_COL, rate_content)

        # H17 账号信息填充
        tag_display = str(current_tag).strip() if pd.notna(current_tag) else ""
        account_content = account + (f" ({current_tag})" if current_tag else "")
        safe_write_cell(summary_sheet, SUMMARY_ACCOUNT_ROW, SUMMARY_ACCOUNT_COL, account_content)

        # ====================== 核心修改：D28-AK50 仅偶数行填充 ======================
        logger.info(f"【{SHEET_SUMMARY}】开始填充12个月明细，仅填充D28-AK50区域偶数行")
        end_month = datetime.strptime(end_period, "%Y-%m")
        # 遍历目标偶数行：idx=0对应28行（13个月前），idx=11对应50行（上个月）
        for idx, current_row in enumerate(SUMMARY_MONTHLY_TARGET_ROWS):
            # 时间逻辑保持不变：最早的月份在最上方，最新的上个月在最下方
            months_ago = 11 - idx
            current_month = end_month - relativedelta(months=months_ago)
            period_str = current_month.strftime("%Y-%m")

            # 获取当月数据
            amount_usd = monthly_summary.get(period_str, 0.0)
            current_exchange_rate = rate_dict.get(period_str, 0.0)
            if current_exchange_rate == 0.0:
                logger.warning(f"{period_str} 未在rate.csv中找到对应汇率，使用0.0计算")

            # 向上取整计算各项金额
            amount_jpy_raw = amount_usd * current_exchange_rate
            amount_jpy = math.ceil(amount_jpy_raw)
            tax_amount = math.ceil(amount_jpy * tax_rate)
            agent_fee = math.ceil((amount_jpy + tax_amount) * agent_rate)
            final_amount = math.ceil(amount_jpy + tax_amount + agent_fee)

            # 写入对应单元格（仅偶数行）
            safe_write_cell(summary_sheet, current_row, MONTHLY_COL_YEAR_MONTH, period_str)
            safe_write_cell(summary_sheet, current_row, MONTHLY_COL_AMOUNT_USD, amount_usd)
            safe_write_cell(summary_sheet, current_row, MONTHLY_COL_AMOUNT_JPY, amount_jpy)
            safe_write_cell(summary_sheet, current_row, MONTHLY_COL_TAX, tax_amount)
            safe_write_cell(summary_sheet, current_row, MONTHLY_COL_AGENT_FEE, agent_fee)
            safe_write_cell(summary_sheet, current_row, MONTHLY_COL_FINAL, final_amount)

            logger.debug(f"填充 {period_str} → 行{current_row}（偶数行）：USD={amount_usd:.2f}, JPY={amount_jpy}")
        # ====================== 偶数行填充结束 ======================

        # ====================== 59行开始的消费明细插入 (已修改：增加边框) ======================
        insert_row = SUMMARY_DETAIL_START_ROW
        if not account_details.empty:
            num_rows = len(account_details)
            summary_sheet.insert_rows(insert_row, amount=num_rows)

            # 定义样式：居中对齐 + 细边框
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # --- 新增：定义边框样式 ---
            thin_side = Side(style='thin', color='000000')  # 黑色细边框
            thin_border = Border(
                left=thin_side,
                right=thin_side,
                top=thin_side,
                bottom=thin_side
            )

            # 继承模板原有金额格式
            try:
                reference_row = SUMMARY_DETAIL_START_ROW - 1
                col_money = SUMMARY_DETAIL_COLS['消费金额']
                ref_cell_money = summary_sheet.cell(row=reference_row, column=col_money)
                money_style = ref_cell_money.number_format
            except Exception:
                money_style = '0.00'

            for row_idx, (_, detail_row) in enumerate(account_details.iterrows(), start=insert_row):
                col_type = SUMMARY_DETAIL_COLS['ProductCode']
                col_id = SUMMARY_DETAIL_COLS['资源id']
                col_money = SUMMARY_DETAIL_COLS['消费金额']

                # 写入数据
                safe_write_cell(summary_sheet, row_idx, col_type, detail_row.get('ProductCode', ''))
                safe_write_cell(summary_sheet, row_idx, col_id, detail_row.get('资源id', ''))
                money_cell = summary_sheet.cell(row=row_idx, column=col_money)
                money_cell.value = detail_row.get('消费金额', 0.0)

                # 合并单元格
                for start_col, end_col in SUMMARY_MERGE_RANGES:
                    try:
                        summary_sheet.merge_cells(start_row=row_idx, start_column=start_col, end_row=row_idx,
                                                  end_column=end_col)
                    except ValueError:
                        pass

                # --- 修改：设置格式与边框 ---
                # 遍历所有合并区域的列，统一设置边框和对齐
                for start_col, end_col in SUMMARY_MERGE_RANGES:
                    for col in range(start_col, end_col + 1):
                        cell = summary_sheet.cell(row=row_idx, column=col)
                        cell.alignment = center_align
                        cell.border = thin_border  # 应用边框

                # 单独设置金额列的数字格式
                money_cell.number_format = money_style

            logger.info(f"【{SHEET_SUMMARY}】消费明细填充完成，共{num_rows}行，已添加边框")
        # ====================== 消费明细插入结束 ======================

    return current_wb


def main():
    logger.info("=" * 80)
    logger.info(f"开始执行课金月报生成程序：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    try:
        # 日期计算
        today = datetime.today()
        target_date = today.replace(day=1) - relativedelta(days=1)
        last_month_year = target_date.year
        last_month = target_date.month

        first_day_last_month = datetime(last_month_year, last_month, 1).strftime("%Y/%m/%d")
        last_day_last_month_dt = datetime(last_month_year, last_month, 1) + relativedelta(months=1) - relativedelta(
            days=1)
        last_day_last_month = last_day_last_month_dt.strftime("%Y/%m/%d")
        date_range = f"{first_day_last_month}～{last_day_last_month}"
        third_day_current_month_dt = today.replace(day=3)
        date_str_d58 = f"{third_day_current_month_dt.year}/{third_day_current_month_dt.month}/{third_day_current_month_dt.day}"

        # 加载数据
        script_dir = os.path.dirname(os.path.abspath(__file__))
        template_path, output_dir, merged_df, aliyun_bills_by_period, start_period, end_period, rate_dict = load_and_preprocess_data(
            script_dir, target_date)

        logger.info(f"处理周期：{start_period} ~ {end_period}")
        logger.info(f"开始生成月报，共 {len(merged_df)} 条数据")

        for idx, row in merged_df.iterrows():
            logger.info(f"\n--- 处理第 {idx + 1}/{len(merged_df)} 条数据 ---")
            account = row['资源所属账号']
            current_tag = row.get('标签')
            project_key = row.get('Project Key', '未知项目')

            # 创建账号-标签文件夹
            folder_name_parts = [account]
            clean_tag = ""
            if current_tag and pd.notna(current_tag):
                tag_str = str(current_tag).strip()
                clean_tag = re.sub(r'[^\w\s-]', '_', tag_str).strip()
                folder_name_parts.append(clean_tag)
            folder_name = "-".join(folder_name_parts)
            output_folder = os.path.join(output_dir, folder_name)
            os.makedirs(output_folder, exist_ok=True)
            logger.info(f"输出文件夹：{output_folder}")

            # 筛选账单数据
            monthly_summary, total_amount, account_details = process_account_bills(account, current_tag,
                                                                                   aliyun_bills_by_period, start_period,
                                                                                   end_period)
            logger.info(f"上月消费：{total_amount:.2f} USD")

            if total_amount == 0.0 and all(v == 0.0 for v in monthly_summary.values()):
                logger.info(f"无消费数据，跳过 {account}")
                continue

            # 填充模板
            current_wb = populate_template(template_path, row, monthly_summary, account_details, date_range,
                                           date_str_d58, rate_dict, end_period, today)

            # 生成日式文件名
            company_name_raw = row.get('客户公司名称', '未知公司')
            company_name_for_file = company_name_raw.replace('   御中', '')
            year_month_str = f"{last_month_year}年{last_month}月"
            output_filename = f"【{company_name_for_file}様向け】アリババクラウド課金代行請求明細_{year_month_str}.xlsx"
            output_path = os.path.join(output_folder, output_filename)

            # 保存文件
            current_wb.save(output_path)
            logger.info(f"✅ 报表生成成功：{output_path}")
            print(
                f"已生成：{output_path}（项目：{project_key}，账号：{account}，标签：{current_tag if current_tag else '无'}）")

        logger.info("\n" + "=" * 80)
        logger.info(f"程序执行完成：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    except Exception as e:
        logger.error(f"程序执行失败：{str(e)}", exc_info=True)


if __name__ == "__main__":
    main()