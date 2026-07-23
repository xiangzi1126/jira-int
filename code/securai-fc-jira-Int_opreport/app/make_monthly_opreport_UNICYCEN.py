# -*- coding: utf-8 -*-
"""ユニ・チャーム様 CENサービス運用報告書 自動生成

模板: 运维月报模板_【ユニ・チャーム株式会社様 】CENサービス運用報告書.xlsx
成品参考: code/jira/26年6月/【ユニ・チャーム株式会社様 】CENサービス運用報告書2026年6月.xlsx

構成 (4シート):
  - 表紙      : 会社名 / 対象期間 / 作成日
  - 総評      : 挨拶 / アラート件数 / 検知内容 / CPU・メモリ使用率表(Proxy001/002)
                / CEN帯域幅6ヶ月表 / 各セクション / アラートBarChart
  - リソース  : LineChart ×5 (Proxy001 CPU/Mem, Proxy002 CPU/Mem, CEN送受信帯域幅)
  - 【補足】推移 : アラート件数(前月/今月) — BarChartのデータ元

データソース:
  - jira_get_alert.csv            (UNICYCEN プロジェクトのアラート)
  - aliyun_op_cmd_summary.csv     (CPU/メモリ使用率 月次集計, Proxy001/002)
  - aliyun_ecs_op.csv             (ECSインスタンス情報: 機種/コア/メモリ)
  - 監控データ_<account>/cpu_*.csv  (CPU時系列, リソースChart用)
  - 監控データ_<account>/memory_*.csv (メモリ時系列, リソースChart用)
  - cen_bandwidth_history.csv     (CEN帯域幅 月次最大値, 6ヶ月表用) ★aliyun_get_cen_metrics.pyが出力
  - 監控データ_<account>/cen_*.csv   (CEN帯域幅時系列, リソースChart用) ★aliyun_get_cen_metrics.pyが出力
"""
import os
import re
import glob
import zipfile
import shutil
import math
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import Series
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText
from openpyxl.chart.plotarea import DataTable
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, RichTextProperties
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import cm_to_EMU
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.styles import Font
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import logging

# 総評シート挿入データのフォント (MS PGothic 10pt)
DATA_FONT = Font(name='MS PGothic', size=10)

# --- I. 常量定义 ---
PROJECT_KEY = 'UNICYCEN'

SHEET_COVER = '表紙'
SHEET_SUMMARY = '総評'
SHEET_RESOURCE = 'リソース '
SHEET_TREND = '【補足】推移'

# CEN テンプレート
TEMPLATE_NAME = '运维月报模板_【ユニ・チャーム株式会社様 】CENサービス運用報告書.xlsx'

# 表紙 行
COVER_COMPANY_ROW = 5
COVER_PERIOD_ROW = 19
COVER_DATE_ROW = 44

# 総評 行 (alert行挿入なし前提の基本位置)
SUMMARY_GREETING_ROW = 1
SUMMARY_OVERVIEW_ROW = 5          # 1、総評 概要 (全体的に安定稼働しております。)
SUMMARY_ALERT_COUNT_ROW = 8        # 2-1アラート件数
SUMMARY_DETECTION_HEADER_ROW = 28  # ■検知内容
SUMMARY_DETECTION_ROW = 29         # 検知内容詳細 (1行目, テンプレ既定文="特にございません。")
SUMMARY_INCIDENT_HEADER_ROW = 31  # ■事象纏め
SUMMARY_INCIDENT_ROW = 32          # 事象纏め詳細 (1行目, テンプレ既定文="特にございません。")
SUMMARY_RESOURCE_HEADER_ROW = 34   # 3、リソース使用状況

# CPU表 (R35=ヘッダ, R36=小ヘッダ, R37-38=データ)
CPU_TABLE_HEADER_ROW = 35
CPU_TABLE_DATA_ROWS = [37, 38]     # Proxy001, Proxy002
# メモリ表 (R39=ヘッダ, R40=小ヘッダ, R41-42=データ)
MEM_TABLE_HEADER_ROW = 39
MEM_TABLE_DATA_ROWS = [41, 42]

# CEN帯域幅 6ヶ月表 (R43-45=1-3月, R46-48=4-6月)
CEN_TABLE1_HEADER_ROW = 43
CEN_TABLE1_ROWS = [44, 45]         # アウトバウンド, インバウンド
CEN_TABLE2_HEADER_ROW = 46
CEN_TABLE2_ROWS = [47, 48]

# セクション見出し行
SECTION_WORK_ROW = 65               # 5、YYYY年M月の作業
SECTION_PLAN_ROW = 71               # 7、YYYY年N月の予定

# リソース sheet Chart 行
RESOURCE_CHART_ROWS = [3, 15, 27, 39, 52]  # Proxy001CPU, Proxy001Mem, Proxy002CPU, Proxy002Mem, CEN
RESOURCE_DATA_START_COL = 16       # P列 (時系列データ退避用)

# 推移 sheet
TREND_DETAIL_HEADER_ROW = 2        # R2: 監視対象/メトリクス名/前月/今月
TREND_DETAIL_START_ROW = 3         # R3-R8
TREND_SUMMARY_HEADER_ROW = 12      # R12: 前月/今月
TREND_SUMMARY_START_ROW = 13       # R13-R14 (Proxy001/002)

# プロキシ表示名マッピング (InstanceName → 表示名)
# JP-Proxy001-20201216 → JP-Proxy001, JP-Proxy002-20201216 → JP-Proxy002
PROXY_HOSTS = ['JP-Proxy001', 'JP-Proxy002']


# --- II. 辅助函数 ---
def init_logger():
    log_dir = os.path.normpath(os.path.join(os.path.dirname(__file__), '../../jira/log'))
    os.makedirs(log_dir, exist_ok=True)
    log_file = os.path.join(log_dir, 'monthly_opreport_UNICYCEN.log')
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler()
        ]
    )
    return logging.getLogger(__name__)


logger = init_logger()


def safe_read_csv(file_path):
    """多编码尝试读取CSV"""
    for encoding in ['utf-8-sig', 'utf-8', 'gbk', 'shift-jis', 'latin1']:
        try:
            return pd.read_csv(file_path, encoding=encoding)
        except (UnicodeDecodeError, UnicodeError):
            continue
        except Exception as e:
            logger.warning(f"读取文件 {file_path} 失败: {e}")
            continue
    raise IOError(f"无法读取文件: {file_path}")


def get_alert_circle_number(n):
    """数字转圆圈数字 ①②③..."""
    circles = '①②③④⑤⑥⑦⑧⑨⑩⑪⑫⑬⑭⑮⑯⑰⑱⑲⑳'
    if 1 <= n <= 20:
        return circles[n - 1]
    return f'({n})'


def _set_chart_size(chart_obj, anchor_str, width_cm, height_cm):
    """设置图表尺寸（OneCellAnchor，保存后用 _fix_chart_sizes_in_xlsx 修正）"""
    col_letter, row = coordinate_from_string(anchor_str)
    col = column_index_from_string(col_letter)
    marker = AnchorMarker(col=col - 1, colOff=0, row=row - 1, rowOff=0)
    size = XDRPositiveSize2D(cx=cm_to_EMU(width_cm), cy=cm_to_EMU(height_cm))
    chart_obj.anchor = OneCellAnchor(_from=marker, ext=size)


def _fix_chart_sizes_in_xlsx(xlsx_path, size_map):
    """保存后修正 xlsx 内 drawing XML 的图表尺寸
    size_map: {drawing序号(1-based): (width_cm, height_cm)}
    """
    tmp_path = xlsx_path + '.tmp'
    pat = re.compile(r'(<ext cx=")([0-9]+)(" cy=")([0-9]+)(")')

    with zipfile.ZipFile(xlsx_path, 'r') as zin:
        with zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                m = re.match(r'xl/drawings/drawing(\d+)\.xml', item.filename)
                if m:
                    drawing_idx = int(m.group(1))
                    if drawing_idx in size_map:
                        w_cm, h_cm = size_map[drawing_idx]
                        new_cx = str(cm_to_EMU(w_cm))
                        new_cy = str(cm_to_EMU(h_cm))
                        text = data.decode('utf-8')
                        text = pat.sub(
                            lambda g: g.group(1) + new_cx + g.group(3) + new_cy + g.group(5),
                            text
                        )
                        data = text.encode('utf-8')
                zout.writestr(item, data)

    os.remove(xlsx_path)
    shutil.move(tmp_path, xlsx_path)
    logger.info(f"图表尺寸修正完成: {size_map}")


def safe_insert_rows(ws, idx, n):
    """insert_rows 安全版。
    openpyxl 3.1.x 的 insert_rows 会平移单元格的值/样式, 但合并单元格只新建平移后的、
    不删除旧的 -> 残留重叠合并(如 Q49:R49 等"错乱")。
    本函数: 插入前 unmerge 插入点及下方(含跨越插入点)的合并, 插入后按平移坐标重新 merge。"""
    if n <= 0:
        return
    # 收集受影响合并的原始坐标 (max_row >= idx: 整体在下方 或 跨越插入点)
    affected = [(mr.min_col, mr.min_row, mr.max_col, mr.max_row)
                for mr in list(ws.merged_cells.ranges) if mr.max_row >= idx]
    # 先 unmerge, 避免 insert_rows 留下旧合并
    for mr in list(ws.merged_cells.ranges):
        if mr.max_row >= idx:
            ws.unmerge_cells(str(mr))
    # 插入行 (openpyxl 会平移值/样式)
    ws.insert_rows(idx, n)
    # 按平移后坐标重新合并
    for c1, r1, c2, r2 in affected:
        new_r1 = r1 + n if r1 >= idx else r1
        new_r2 = r2 + n
        ws.merge_cells(start_row=new_r1, start_column=c1, end_row=new_r2, end_column=c2)


import calendar


def days_in_month(year, month):
    return calendar.monthrange(year, month)[1]


# --- III. 数据加载 ---
def load_all_data(script_dir):
    """加载所有数据源"""
    def get_path(rel_path):
        return os.path.normpath(os.path.join(script_dir, rel_path))

    data_dir = get_path("../../jira/data")
    output_dir = get_path("../../jira/monthly_report")
    template_path = get_path(f"../../jira/monthly_report/{TEMPLATE_NAME}")
    os.makedirs(output_dir, exist_ok=True)

    # 必須データ
    op_account_csv = os.path.join(data_dir, 'jira_get_op_account.csv')
    ecs_csv = os.path.join(data_dir, 'aliyun_ecs_op.csv')
    cmd_summary_csv = os.path.join(data_dir, 'aliyun_op_cmd_summary.csv')
    alert_csv = os.path.join(data_dir, 'jira_get_alert.csv')

    for path in [op_account_csv, ecs_csv, cmd_summary_csv, alert_csv]:
        if not os.path.exists(path):
            raise FileNotFoundError(f"关键文件不存在: {path}")
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"模板文件不存在: {template_path}")

    op_account_df = safe_read_csv(op_account_csv)
    ecs_df = safe_read_csv(ecs_csv)
    cmd_summary_df = safe_read_csv(cmd_summary_csv)
    alert_df = safe_read_csv(alert_csv)

    # CEN帯域幅履歴 (オプション — 未生成時は空)
    cen_history_csv = os.path.join(data_dir, 'cen_bandwidth_history.csv')
    cen_history_df = safe_read_csv(cen_history_csv) if os.path.exists(cen_history_csv) else pd.DataFrame()
    if cen_history_df.empty:
        logger.warning(f"CEN帯域幅履歴ファイル未検出: {cen_history_csv} (CEN帯域幅表は空欄になります)")

    # 名称マッピング
    names_csv = get_path("../../jira/config/opreport_names.csv")
    names_df = safe_read_csv(names_csv) if os.path.exists(names_csv) else pd.DataFrame()

    logger.info(f"数据加载: ECS{len(ecs_df)}台, 资源汇总{len(cmd_summary_df)}条, "
                f"告警{len(alert_df)}条, CEN历史{len(cen_history_df)}条")
    return {
        'output_dir': output_dir,
        'template_path': template_path,
        'data_dir': data_dir,
        'op_account': op_account_df,
        'ecs': ecs_df,
        'cmd_summary': cmd_summary_df,
        'alert': alert_df,
        'cen_history': cen_history_df,
        'names': names_df,
    }


def get_proxy_cmd_data(cmd_df, account):
    """アカウント配下のProxy001/002のリソース集計を取得"""
    proj = cmd_df[cmd_df['资源所属账号'] == account].copy()
    result = {}
    for short_name in PROXY_HOSTS:
        # JP-Proxy001-20201216 等を前方一致で取得
        mask = proj['InstanceName'].astype(str).str.startswith(short_name, na=False)
        matched = proj[mask]
        if not matched.empty:
            result[short_name] = matched.iloc[0]
        else:
            result[short_name] = None
    return result


def get_proxy_ecs_info(ecs_df, account):
    """Proxy001/002のECS情報 (機種/コア/メモリ)を取得"""
    proj = ecs_df[ecs_df['资源所属账号'] == account].copy()
    result = {}
    for short_name in PROXY_HOSTS:
        mask = proj['InstanceName'].astype(str).str.startswith(short_name, na=False)
        matched = proj[mask]
        if not matched.empty:
            row = matched.iloc[0]
            result[short_name] = {
                'instance_name': str(row['InstanceName']),
                'instance_id': str(row['InstanceId']),
                'cpu_cores': row['CPU(核)'],
                'memory_gb': round(float(row['Memory(MB)']) / 1024),
            }
        else:
            result[short_name] = None
    return result


def get_alerts_for_project(alert_df, project_key, target_year, target_month):
    """UNICYCENの今月・前月アラート (备注='不记录月报' を除外)"""
    proj_alerts = alert_df[alert_df['Project Key'] == project_key].copy()
    if proj_alerts.empty:
        return pd.DataFrame(), pd.DataFrame()

    if '备注' in proj_alerts.columns:
        proj_alerts = proj_alerts[
            proj_alerts['备注'].astype(str).str.strip() != '不记录月报'
        ]
    if proj_alerts.empty:
        return pd.DataFrame(), pd.DataFrame()

    # JST変換後タイムゾーン除去
    proj_alerts['告警检测日时'] = pd.to_datetime(
        proj_alerts['告警检测日时'], errors='coerce', utc=True
    ).dt.tz_convert('Asia/Tokyo').dt.tz_localize(None)
    proj_alerts = proj_alerts.dropna(subset=['告警检测日时'])

    current_start = datetime(target_year, target_month, 1)
    current_end = current_start + relativedelta(months=1) - timedelta(seconds=1)
    prev_start = current_start - relativedelta(months=1)
    prev_end = current_start - timedelta(seconds=1)

    current_alerts = proj_alerts[
        (proj_alerts['告警检测日时'] >= current_start) &
        (proj_alerts['告警检测日时'] <= current_end)
    ].sort_values('告警检测日时')
    prev_alerts = proj_alerts[
        (proj_alerts['告警检测日时'] >= prev_start) &
        (proj_alerts['告警检测日时'] <= prev_end)
    ]
    return current_alerts, prev_alerts


def load_timeseries(data_dir, account, instance_name, prefix):
    """監視時系列CSV読込 (cpu_/memory_/cen_)"""
    monitor_dir = os.path.join(data_dir, f'监控数据_{account}')
    if not os.path.exists(monitor_dir):
        return pd.DataFrame()
    pattern = os.path.join(monitor_dir, f'{prefix}_*.csv')
    files = glob.glob(pattern)
    for f in files:
        basename = os.path.basename(f)
        name_part = basename.replace(f'{prefix}_', '').replace('.csv', '')
        if name_part.lower() == instance_name.lower():
            try:
                df = safe_read_csv(f)
                if 'timestamp' in df.columns:
                    df['datetime'] = pd.to_datetime(df['timestamp'], unit='ms')
                    df = df.sort_values('datetime')
                return df
            except Exception as e:
                logger.warning(f"读取监控数据失败 {f}: {e}")
                return pd.DataFrame()
    return pd.DataFrame()


def load_cen_timeseries(data_dir, account):
    """CEN帯域幅時系列CSV読込 (cen_<name>.csv)"""
    monitor_dir = os.path.join(data_dir, f'监控数据_{account}')
    if not os.path.exists(monitor_dir):
        return pd.DataFrame()
    files = glob.glob(os.path.join(monitor_dir, 'cen_*.csv'))
    if not files:
        return pd.DataFrame()
    try:
        df = safe_read_csv(files[0])
        if 'timestamp' in df.columns:
            df['datetime'] = pd.to_datetime(df['timestamp'], unit='ms')
            df = df.sort_values('datetime')
        return df
    except Exception as e:
        logger.warning(f"CEN時系列読込失敗 {files[0]}: {e}")
        return pd.DataFrame()


def safe_float(val):
    """N/A / 空 → None, それ以外 float"""
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return None
    s = str(val).strip()
    if s == '' or s.upper() in ('NAN', 'N/A', 'NONE'):
        return None
    try:
        return float(s)
    except ValueError:
        return None


# --- IV. Sheet 填充函数 ---
def populate_cover(wb, company_name, target_year, target_month, today):
    """表紙"""
    if SHEET_COVER not in wb.sheetnames:
        return
    ws = wb[SHEET_COVER]
    ws.cell(row=COVER_COMPANY_ROW, column=1).value = f"{company_name} 御中"

    last_day = days_in_month(target_year, target_month)
    ws.cell(row=COVER_PERIOD_ROW, column=1).value = (
        f"{target_year}/{target_month}/{1}～{target_year}/{target_month}/{last_day}"
    )
    ws.cell(row=COVER_DATE_ROW, column=1).value = f"作成日：{today.year}/{today.month}/{today.day}"
    logger.info(f"【表紙】填充完成: {company_name}, {target_year}年{target_month}月")


def populate_summary(wb, current_alerts, prev_alerts, proxy_cmd, proxy_ecs,
                     cen_history, target_year, target_month):
    """総評 (アラート行挿入対応)
    構成: 1、総評概要(备注) → ■検知内容 → ■事象纏め → 3、リソース使用状況 → ...
    make_monthly_opreport.py のパターンを踏襲。
    """
    if SHEET_SUMMARY not in wb.sheetnames:
        return
    ws = wb[SHEET_SUMMARY]

    def set_cell(row, col, value):
        """セルに値を書き込み + フォント(MS PGothic 10pt)を設定"""
        c = ws.cell(row=row, column=col)
        c.value = value
        c.font = DATA_FONT

    # 挨拶文
    set_cell(SUMMARY_GREETING_ROW, 1, (
        f"平素より弊社運用管理代行サービスをご利用いただき、誠にありがとうございます。\n"
        f"{target_year}年{target_month}月のCENサービス運用状況について、下記の通り、ご報告申し上げます。"
    ))

    alert_count = len(current_alerts)
    offset = 0  # 行挿入による累積オフセット

    # === 1、総評 概要 (R5) ===
    # 0件: テンプレ既定文 "全体的に安定稼働しております。" を維持
    # N件: R5をクリア → R6に(N-1)行挿入 → 各告警の备注を "M/D 、<备注>" で出力
    if alert_count > 0:
        set_cell(SUMMARY_OVERVIEW_ROW, 1, None)  # 既定文クリア
        if alert_count > 1:
            insert_at = SUMMARY_OVERVIEW_ROW + 1   # R6
            insert_n = alert_count - 1
            safe_insert_rows(ws, insert_at, insert_n)
            offset += insert_n
            logger.info(f"【総評】総評概要に{insert_n}行挿入")

        for idx, (_, row) in enumerate(current_alerts.iterrows()):
            r = SUMMARY_OVERVIEW_ROW + idx
            detect_time = row['告警检测日时']
            remark = row.get('备注', '')
            if pd.isna(remark) or not str(remark).strip():
                remark = f"{row.get('告警目标', '')}サーバのアラートを検知しました。"
            date_str = f"{detect_time.month}/{detect_time.day}"
            set_cell(r, 1, f"{date_str} 、{str(remark).strip()}")

    # 2-1アラート件数 (offset適用済み行に書き込み)
    set_cell(SUMMARY_ALERT_COUNT_ROW + offset, 1, f"2-1アラート件数{alert_count}件")

    # === ■検知内容 (R28+offset header, R29+offset data slot) ===
    detection_data_start = SUMMARY_DETECTION_ROW + offset  # 挿入前のデータ開始行
    if alert_count <= 0:
        set_cell(detection_data_start, 1, "特にございません。")
    else:
        set_cell(detection_data_start, 1, None)
        if alert_count > 1:
            insert_at = detection_data_start + 1   # R30+offset
            insert_n = alert_count - 1
            safe_insert_rows(ws, insert_at, insert_n)
            offset += insert_n
            logger.info(f"【総評】検知内容に{insert_n}行挿入")

        for idx, (_, row) in enumerate(current_alerts.iterrows()):
            r = detection_data_start + idx
            detect_time = row['告警检测日时']
            alert_desc = str(row.get('告警简述', '') or '').strip()
            date_str = (f"{detect_time.year}-{detect_time.month}-{detect_time.day} "
                        f"{detect_time.hour}:{detect_time.minute:02d}")
            circle_num = get_alert_circle_number(idx + 1)
            text = f"アラート{circle_num}：{date_str} {alert_desc}".strip()
            set_cell(r, 1, text)

    # === ■事象纏め (R31+offset header, R32+offset data slot) ===
    incident_data_row = SUMMARY_INCIDENT_ROW + offset

    reason_groups = {}
    if alert_count > 0:
        for idx, (_, row) in enumerate(current_alerts.iterrows()):
            reason = row.get('告警原因', '')
            if pd.isna(reason) or not str(reason).strip():
                continue
            reason = str(reason).strip()
            if reason not in reason_groups:
                reason_groups[reason] = []
            reason_groups[reason].append(idx + 1)

    if alert_count <= 0 or not reason_groups:
        set_cell(incident_data_row, 1, "特にございません。")
    else:
        set_cell(incident_data_row, 1, None)
        reasons = list(reason_groups.items())
        if len(reasons) > 1:
            insert_at = incident_data_row + 1
            insert_n = len(reasons) - 1
            safe_insert_rows(ws, insert_at, insert_n)
            offset += insert_n
            logger.info(f"【総評】事象纏めに{insert_n}行挿入")

        for ridx, (reason, alert_nums) in enumerate(reasons):
            r = incident_data_row + ridx
            nums_str = '、'.join([get_alert_circle_number(n) for n in alert_nums])
            text = f"アラート{nums_str}：{reason}"
            set_cell(r, 1, text)

    # === 以降の行に offset を適用 ===
    _fill_cpu_table(ws, proxy_cmd, proxy_ecs, offset)
    _fill_memory_table(ws, proxy_cmd, proxy_ecs, offset)
    # CEN帯域幅6ヶ月表: 純模板(合并+ラベルは safe_insert_rows で平移保留, データは手動入力)
    _fill_section_dates(ws, target_year, target_month, offset)

    # BarChart (推移シート参照) — アラート件数の右下
    chart_row = SUMMARY_ALERT_COUNT_ROW + 1 + offset
    _create_alert_chart(wb, ws, f'B{chart_row}')

    logger.info(f"【総評】填充完成 (alert={alert_count}件, reasons={len(reason_groups)}, offset={offset})")


def _fill_cpu_table(ws, proxy_cmd, proxy_ecs, off):
    """CPU使用率表 - 定位填入模板(保留合并): C列搜ホスト名定位行, 只写 E/F/G/H 数据。
    模板合并(B35:B36機種, C35:D36ホスト名, E35:H35表头, B37:B38機種, C37:D38ホスト名)
    及表头标签由 safe_insert_rows 正确平移保留, 不再 unmerge/重建(清空重建会丢合并)。"""
    cpu_start = CPU_TABLE_HEADER_ROW + off
    cpu_end = CPU_TABLE_HEADER_ROW + 3 + off
    for short_name in PROXY_HOSTS:
        found_row = None
        for r in range(cpu_start, cpu_end + 1):
            cv = ws.cell(row=r, column=3).value
            if cv and str(cv).strip() == short_name:
                found_row = r
                break
        if found_row is None:
            logger.warning(f"【総評】CPU表: C列に {short_name} が見つかりません")
            continue

        ecs = proxy_ecs.get(short_name)
        cmd = proxy_cmd.get(short_name)
        cores = ecs['cpu_cores'] if ecs else ''
        ws.cell(row=found_row, column=5).value = cores           # E: コア数
        if cmd is not None:
            ws.cell(row=found_row, column=6).value = safe_float(cmd.get('CPU最小使用率(%)'))   # F: 最小
            ws.cell(row=found_row, column=7).value = safe_float(cmd.get('CPU平均使用率(%)'))   # G: 平均
            ws.cell(row=found_row, column=8).value = safe_float(cmd.get('CPU最大使用率(%)'))   # H: 最大
    logger.info(f"【総評】CPU表填充 (offset={off}, 定位填入/保留合并)")


def _fill_memory_table(ws, proxy_cmd, proxy_ecs, off):
    """メモリ使用率表 - 定位填入模板(保留合并): C列搜ホスト名定位行, 只写 E/F/G/H 数据。
    模板合并(B39:B40機種, C39:D40ホスト名, E39:H39表头, B41:B42機種, C41:D42ホスト名)
    及表头标签由 safe_insert_rows 正确平移保留, 不再 unmerge/重建。"""
    mem_start = MEM_TABLE_HEADER_ROW + off
    mem_end = MEM_TABLE_HEADER_ROW + 3 + off
    for short_name in PROXY_HOSTS:
        found_row = None
        for r in range(mem_start, mem_end + 1):
            cv = ws.cell(row=r, column=3).value
            if cv and str(cv).strip() == short_name:
                found_row = r
                break
        if found_row is None:
            logger.warning(f"【総評】メモリ表: C列に {short_name} が見つかりません")
            continue

        ecs = proxy_ecs.get(short_name)
        cmd = proxy_cmd.get(short_name)
        mem_gb = ecs['memory_gb'] if ecs else ''
        ws.cell(row=found_row, column=5).value = mem_gb          # E: メモリサイズ(GB)
        if cmd is not None:
            ws.cell(row=found_row, column=6).value = safe_float(cmd.get('内存最小使用率(%)'))  # F
            ws.cell(row=found_row, column=7).value = safe_float(cmd.get('内存平均使用率(%)'))  # G
            ws.cell(row=found_row, column=8).value = safe_float(cmd.get('内存最大使用率(%)'))  # H
    logger.info(f"【総評】メモリ表填充 (offset={off}, 定位填入/保留合并)")


def _fill_section_dates(ws, target_year, target_month, off):
    """セクション5/7 の年月プレースホルダ置換"""
    next_month_date = datetime(target_year, target_month, 1) + relativedelta(months=1)
    next_year = next_month_date.year
    next_month = next_month_date.month

    # 5、YYYY年M月の作業
    ws.cell(row=SECTION_WORK_ROW + off, column=1).value = f"5、{target_year}年{target_month}月の作業"
    # 7、YYYY年N月の予定
    ws.cell(row=SECTION_PLAN_ROW + off, column=1).value = f"7、{next_year}年{next_month}月の予定"


def _create_alert_chart(wb, summary_ws, anchor_cell):
    """推移シート参照のアラート件数 BarChart"""
    if SHEET_TREND not in wb.sheetnames:
        return
    trend_ws = wb[SHEET_TREND]

    # R12=前月/今月 ヘッダ, R13-14=Proxy001/002
    data_ref = Reference(trend_ws, min_col=3, min_row=TREND_SUMMARY_HEADER_ROW,
                         max_col=4, max_row=TREND_SUMMARY_START_ROW + 1)
    cats_ref = Reference(trend_ws, min_col=1, min_row=TREND_SUMMARY_START_ROW,
                         max_row=TREND_SUMMARY_START_ROW + 1)

    chart = BarChart()
    chart.type = "col"
    chart.title = "対応アラート件数"
    chart.style = 10
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 10
    chart.y_axis.tickLblPos = 'nextTo'
    chart.y_axis.delete = False
    chart.x_axis.tickLblPos = 'low'
    chart.x_axis.delete = False
    chart.x_axis.txPr = RichText(
        bodyPr=RichTextProperties(rot=0),
        p=[Paragraph(
            pPr=ParagraphProperties(defRPr=CharacterProperties(sz=900)),
            endParaRPr=CharacterProperties(sz=900)
        )]
    )
    chart.legend = None
    chart.plot_area.dTable = DataTable(showHorzBorder=True, showVertBorder=True, showKeys=True)

    # 绘画区留白(UNICYCEN 特例, 非通用): 顶部 15%, 底部 7% (y=0.15, h=0.78; 与リソース图/TSIINFRA 不同)
    from openpyxl.chart.layout import Layout, ManualLayout
    chart.layout = Layout(
        manualLayout=ManualLayout(
            layoutTarget='inner',
            xMode='edge', yMode='edge',
            wMode='edge', hMode='edge',
            x=0.08, y=0.15, w=0.90, h=0.78,
        )
    )

    summary_ws.add_chart(chart, anchor_cell)
    _set_chart_size(summary_ws._charts[-1], anchor_cell, 15, 8)
    logger.info(f"【総評】BarChart作成 at {anchor_cell}")


def populate_trend_sheet(wb, current_alerts, prev_alerts):
    """【補足】推移 — アラート件数(前月/今月)"""
    if SHEET_TREND not in wb.sheetnames:
        return
    ws = wb[SHEET_TREND]

    # 推移シートの監視対象名は短縮名 (Proxy001/Proxy002)
    trend_names = ['Proxy001', 'Proxy002']

    def count_alerts(alerts, proxy_name):
        if alerts is None or alerts.empty:
            return 0
        return len(alerts[alerts['告警目标'].astype(str).str.lower().str.contains(
            proxy_name.lower(), na=False)])

    # 詳細表 (R3-R8): Proxy001/002 × CPU/Memory/Agent_heartbeat
    detail_rows = [
        (3, 'Proxy001', 'CPUUtilization'),
        (4, None, 'MemoryUtilization'),
        (5, None, 'Agent_no_heartbeat'),
        (6, 'Proxy002', 'CPUUtilization'),
        (7, None, 'MemoryUtilization'),
        (8, None, 'Agent_no_heartbeat'),
    ]
    for r, proxy, metric in detail_rows:
        if proxy:
            ws.cell(row=r, column=1).value = proxy
            ws.cell(row=r, column=2).value = metric
            ws.cell(row=r, column=3).value = count_alerts(prev_alerts, proxy)
            ws.cell(row=r, column=4).value = count_alerts(current_alerts, proxy)
        else:
            ws.cell(row=r, column=2).value = metric
            ws.cell(row=r, column=3).value = 0
            ws.cell(row=r, column=4).value = 0

    # サマリ表 (R12-R14): BarChartのデータ元
    ws.cell(row=TREND_SUMMARY_HEADER_ROW, column=3).value = '前月'
    ws.cell(row=TREND_SUMMARY_HEADER_ROW, column=4).value = '今月'
    for i, proxy in enumerate(trend_names):
        r = TREND_SUMMARY_START_ROW + i
        ws.cell(row=r, column=1).value = proxy
        ws.cell(row=r, column=3).value = count_alerts(prev_alerts, proxy)
        ws.cell(row=r, column=4).value = count_alerts(current_alerts, proxy)

    logger.info(f"【推移】填充完成")


def populate_resource_sheet(wb, proxy_ecs, account, data_dir):
    """リソース sheet — LineChart ×5"""
    if SHEET_RESOURCE not in wb.sheetnames:
        return
    ws = wb[SHEET_RESOURCE]

    # (title_row, instance_fullname, metric_label, prefix)
    charts_def = []
    for i, short_name in enumerate(PROXY_HOSTS):
        ecs = proxy_ecs.get(short_name)
        fullname = ecs['instance_name'] if ecs else short_name
        charts_def.append((RESOURCE_CHART_ROWS[i * 2], fullname, 'CPU', 'cpu'))
        charts_def.append((RESOURCE_CHART_ROWS[i * 2 + 1], fullname, 'メモリ', 'memory'))
    # CEN送受信帯域幅
    charts_def.append((RESOURCE_CHART_ROWS[4], None, 'CEN送信/受信', 'cen'))

    for idx, (title_row, fullname, label, prefix) in enumerate(charts_def):
        data_col = RESOURCE_DATA_START_COL + idx * 3   # P, S, V, Y, AB

        # タイトル
        if prefix == 'cen':
            ws.cell(row=title_row, column=2).value = "■CEN送信/受信帯域幅"
            df = load_cen_timeseries(data_dir, account)
        else:
            ws.cell(row=title_row, column=2).value = f"■{fullname}({label}使用率)"
            df = load_timeseries(data_dir, account, fullname, prefix)

        if df is None or df.empty or 'datetime' not in df.columns:
            logger.warning(f"  [リソース] {label} {fullname or 'CEN'}: 監視データなし → チャート省略")
            continue

        _create_resource_linechart(ws, df, title_row, data_col, label, fullname, prefix)
        logger.info(f"  [リソース] {label} {fullname or 'CEN'}: {len(df)}ポイント → Chart作成")


def _create_resource_linechart(ws, df, title_row, data_col, label, fullname, prefix):
    """単一LineChart作成 (CENは2系列)"""
    header_row = title_row
    num_points = len(df)

    # 時系列データ退避 (data_col〜)
    ws.cell(row=header_row, column=data_col).value = 'datetime'
    if prefix == 'cen':
        # 送信/受信 2系列
        send_col = df.columns[[c.lower() in ('送信(mbps)', '送信', '送信(mbps)', 'outbound', 'out', '送信(mbps)') for c in df.columns]]
        # 柔軟に列名探索
        send_name = next((c for c in df.columns if '送信' in str(c) or 'out' in str(c).lower()), None)
        recv_name = next((c for c in df.columns if '受信' in str(c) or ('in' in str(c).lower() and '年' not in str(c).lower() and 'min' not in str(c).lower())), None)
        ws.cell(row=header_row, column=data_col + 1).value = '送信(Mbps)'
        ws.cell(row=header_row, column=data_col + 2).value = '受信(Mbps)'
        for pt_idx, (_, pt) in enumerate(df.iterrows()):
            r = header_row + 1 + pt_idx
            ws.cell(row=r, column=data_col).value = pt['datetime'].strftime('%m/%d %H:%M')
            ws.cell(row=r, column=data_col + 1).value = pt.get(send_name, 0) if send_name else 0
            ws.cell(row=r, column=data_col + 2).value = pt.get(recv_name, 0) if recv_name else 0
        num_series = 2
        series_cols = [data_col + 1, data_col + 2]
    else:
        # CPU/Memory 1系列 (average)
        ws.cell(row=header_row, column=data_col + 1).value = 'average'
        for pt_idx, (_, pt) in enumerate(df.iterrows()):
            r = header_row + 1 + pt_idx
            ws.cell(row=r, column=data_col).value = pt['datetime'].strftime('%m/%d %H:%M')
            ws.cell(row=r, column=data_col + 1).value = pt.get('average', 0)
        num_series = 1
        series_cols = [data_col + 1]

    chart = LineChart()
    chart.title = (f"{fullname} ({label}使用率)" if prefix != 'cen'
                   else "CEN送信/受信帯域幅")
    chart.style = 10
    chart.y_axis.title = 'Mbps' if prefix == 'cen' else '%'
    chart.x_axis.title = '日時'
    chart.y_axis.tickLblPos = 'nextTo'
    chart.x_axis.tickLblPos = 'low'
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.legend = None if num_series == 1 else True

    # データ参照
    for sc in series_cols:
        data_ref = Reference(ws, min_col=sc, min_row=header_row,
                             max_col=sc, max_row=header_row + num_points)
        chart.add_data(data_ref, titles_from_data=True)
    cats_ref = Reference(ws, min_col=data_col, min_row=header_row + 1,
                         max_row=header_row + num_points)
    chart.set_categories(cats_ref)

    # 線のスタイル (1系列時は青)
    if num_series == 1:
        s = chart.series[0]
        s.graphicalProperties = GraphicalProperties()
        s.graphicalProperties.line = LineProperties(solidFill="4472C4", w=12700)
    else:
        colors = ["4472C4", "ED7D31"]  # 送信=青, 受信=橙
        for i, s in enumerate(chart.series):
            s.graphicalProperties = GraphicalProperties()
            s.graphicalProperties.line = LineProperties(solidFill=colors[i % 2], w=12700)
            chart.legend = True

    anchor = f"B{title_row + 1}"
    ws.add_chart(chart, anchor)
    _set_chart_size(ws._charts[-1], anchor, 12, 6)


# --- V. 主逻辑 ---
def build_project_report(data, target_year, target_month, today):
    """UNICYCEN CENサービス運用報告書 生成"""
    project_key = PROJECT_KEY

    # 名称マッピング
    names_df = data.get('names', pd.DataFrame())
    name_row = (names_df[names_df['Project Key'] == project_key]
                if not names_df.empty else pd.DataFrame())
    if not name_row.empty:
        company_name = str(name_row.iloc[0]['公司名'])
        system_name = str(name_row.iloc[0]['系统名'])
    else:
        logger.warning(f"项目 {project_key} 无名称映射，使用默认")
        company_name = 'ユニ・チャーム株式会社'
        system_name = 'CENサービス'

    logger.info(f"\n{'='*60}")
    logger.info(f"开始处理: {project_key} ({company_name}) {target_year}年{target_month}月")
    logger.info(f"{'='*60}")

    # アカウント
    account_df = data['op_account']
    proj_accounts = account_df[account_df['Project Key'] == project_key]
    if proj_accounts.empty:
        logger.warning(f"项目 {project_key} 无对应账号，跳过")
        return None
    account = proj_accounts.iloc[0]['资源所属账号']
    logger.info(f"账号: {account}")

    # Proxy情報
    proxy_ecs = get_proxy_ecs_info(data['ecs'], account)
    proxy_cmd = get_proxy_cmd_data(data['cmd_summary'], account)
    for sn in PROXY_HOSTS:
        if proxy_ecs.get(sn):
            logger.info(f"  {sn}: {proxy_ecs[sn]['instance_name']} "
                        f"({proxy_ecs[sn]['cpu_cores']}Core/{proxy_ecs[sn]['memory_gb']}GB)")
        else:
            logger.warning(f"  {sn}: ECS情報なし")
        if proxy_cmd.get(sn) is None:
            logger.warning(f"  {sn}: リソース集計(cmd_summary)なし")

    # アラート
    current_alerts, prev_alerts = get_alerts_for_project(
        data['alert'], project_key, target_year, target_month
    )
    logger.info(f"告警(过滤后): 今月{len(current_alerts)}件, 前月{len(prev_alerts)}件")

    # テンプレート読込
    wb = load_workbook(data['template_path'])
    logger.info(f"使用模板: {os.path.basename(data['template_path'])}")

    # 各Sheet填充
    populate_cover(wb, company_name, target_year, target_month, today)
    populate_summary(wb, current_alerts, prev_alerts, proxy_cmd, proxy_ecs,
                     data['cen_history'], target_year, target_month)
    populate_trend_sheet(wb, current_alerts, prev_alerts)
    # リソース sheet は模板のまま変更しない (チャート/データは手動入力)

    # 保存
    output_dir = data['output_dir']
    project_folder = os.path.join(output_dir, project_key)
    os.makedirs(project_folder, exist_ok=True)

    year_month_str = f"{target_year}{target_month:02d}"
    output_filename = f"【{company_name}様 】CENサービス運用報告書{target_year}年{target_month}月.xlsx"
    output_path = os.path.join(project_folder, output_filename)

    wb.save(output_path)
    wb.close()

    # 図表サイズ修正 (drawing1=総評BarChart のみ, リソースsheetは模板維持で変更なし)
    _fix_chart_sizes_in_xlsx(output_path, {
        1: (15, 8),   # 総評 BarChart: 15cm × 8cm (width+2, height-1)
    })

    logger.info(f"报表生成成功: {output_path}")
    return output_path


def main():
    logger.info("=" * 80)
    logger.info(f"开始执行 UNICYCEN CEN運用報告書生成: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    try:
        today = datetime.today()
        target_date = today.replace(day=1) - timedelta(days=1)
        target_year = target_date.year
        target_month = target_date.month
        logger.info(f"目标月份: {target_year}年{target_month}月")

        script_dir = os.path.dirname(os.path.abspath(__file__))
        data = load_all_data(script_dir)

        result = build_project_report(data, target_year, target_month, today)
        if result:
            logger.info(f"\n{'='*80}")
            logger.info(f"UNICYCEN 报表生成成功: {result}")
        else:
            logger.error("UNICYCEN 报表生成失败")

        logger.info(f"结束时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    except Exception as e:
        logger.error(f"程序执行失败: {str(e)}", exc_info=True)


if __name__ == "__main__":
    main()
